"""WP4-A: local file import -> immutable raw parquet snapshot + manifest.

Covers the generic import primitive (text lines / CSV), source resolution,
manifest provenance, single-snapshot re-run semantics, and the
``compsval ingest file`` CLI surface.
"""

from datetime import UTC, datetime
from pathlib import Path

import pytest

from compsval import cli
from compsval.catalog import list_snapshots
from compsval.ingest.import_file import (
    file_to_table,
    import_local_file,
    resolve_source_dir,
)
from compsval.ingest.manifests import read_manifest
from compsval.ingest.snapshots import DATA_FILENAME


def test_file_to_table_txt_preserves_lines(tmp_path: Path) -> None:
    path = tmp_path / "sample.txt"
    path.write_text("a\nb\n\n", encoding="utf-8")
    table = file_to_table(path)
    # trailing newline does not create an empty trailing row; interior blank kept
    assert table.column_names == ["line_no", "content"]
    assert table.column("content").to_pylist() == ["a", "b", ""]
    assert table.column("line_no").to_pylist() == [1, 2, 3]


def test_file_to_table_csv_is_parsed(tmp_path: Path) -> None:
    path = tmp_path / "sample.csv"
    path.write_text("id,price\n1,3200000\n2,4500000\n", encoding="utf-8")
    table = file_to_table(path)
    assert table.column_names == ["id", "price"]
    assert table.num_rows == 2


def test_file_to_table_xlsx_reads_structurally(tmp_path: Path) -> None:
    """EXTFP0-C：XLSX 经 openpyxl 只读结构读取，不伪装为 UTF-8 文本。"""
    from openpyxl import Workbook

    path = tmp_path / "成交.xlsx"
    wb = Workbook()
    ws = wb.create_sheet("成交")
    ws["A1"] = "小区"
    ws["A2"] = 123
    ws["B2"] = True
    ws["C1"] = 3.5
    wb.save(path)

    table = file_to_table(path)
    assert table.column_names == ["sheet_name", "row", "column", "value", "value_type"]
    cells = {
        (row, col): (val, typ)
        for row, col, val, typ in zip(
            table.column("row").to_pylist(),
            table.column("column").to_pylist(),
            table.column("value").to_pylist(),
            table.column("value_type").to_pylist(),
            strict=True,
        )
    }
    assert cells[(1, "A")] == ("小区", "string")
    assert cells[(2, "A")] == ("123", "number")
    assert cells[(2, "B")] == ("True", "bool")
    assert cells[(1, "C")] == ("3.5", "number")
    assert all(n == "成交" for n in table.column("sheet_name").to_pylist())


def test_import_local_file_xlsx_produces_snapshot(tmp_path: Path) -> None:
    """EXTFP0-C：XLSX 可经 import_local_file 形成原始快照，不触发文本读取失败。"""
    from openpyxl import Workbook

    path = tmp_path / "raw.xlsx"
    wb = Workbook()
    ws = wb.create_sheet("Sheet1")
    ws["A1"] = "h1"
    ws["A2"] = "v1"
    ws["B2"] = 7
    wb.save(path)

    fetched_at = datetime(2026, 8, 24, 0, 0, 0, tzinfo=UTC)
    result = import_local_file(
        input_path=path,
        source="SRC-011",
        dataset="chengjiao_xlsx",
        fetched_at=fetched_at,
        query=str(path),
        data_dir=tmp_path / "lake",
    )
    manifest = read_manifest(result.directory)
    assert manifest.source == "lianjia_ext"
    assert manifest.row_count == 3  # 三个已填充单元格
    (ref,) = list_snapshots(tmp_path / "lake")
    assert ref.dataset == "chengjiao_xlsx"
    assert ref.source == "lianjia_ext"


@pytest.mark.parametrize(
    ("token", "expected"),
    [("SRC-007", "lianjia"), ("lianjia", "lianjia"), ("SRC-005", "fang_esf")],
)
def test_resolve_source_dir(token: str, expected: str) -> None:
    assert resolve_source_dir(token) == expected


def test_resolve_source_dir_unknown_raises() -> None:
    with pytest.raises(ValueError):
        resolve_source_dir("not-a-source")


def test_import_local_file_builds_manifest_and_catalog_visible(tmp_path: Path) -> None:
    raw = tmp_path / "evidence" / "lianjia.txt"
    raw.parent.mkdir(parents=True)
    raw.write_text("示例小区126 2室1厅 84.04平米\n挂牌258万成交周期89天\n", encoding="utf-8")

    fetched_at = datetime(2026, 8, 21, 3, 4, 5, tzinfo=UTC)
    result = import_local_file(
        input_path=raw,
        source="SRC-007",
        dataset="chengjiao_list",
        fetched_at=fetched_at,
        query="https://lianjia.com.example/chengjiao/targetdistrict/",
        data_dir=tmp_path / "lake",
    )

    manifest = read_manifest(result.directory)
    assert manifest.source == "lianjia"
    assert manifest.dataset == "chengjiao_list"
    assert manifest.row_count == 2
    assert manifest.source_row_count == 2
    assert manifest.files[0].path == DATA_FILENAME
    assert len(manifest.files[0].sha256) == 64  # sha256 hex

    (ref,) = list_snapshots(tmp_path / "lake")
    assert ref.source == "lianjia"
    assert ref.dataset == "chengjiao_list"


def test_import_server_reimport_is_single_snapshot(tmp_path: Path) -> None:
    raw = tmp_path / "e.txt"
    raw.write_text("x\n", encoding="utf-8")
    fetched_at = datetime(2026, 8, 21, 3, 0, 0, tzinfo=UTC)
    import_local_file(
        input_path=raw,
        source="lianjia",
        dataset="d",
        fetched_at=fetched_at,
        query="q",
        data_dir=tmp_path / "lake",
    )
    # identical import at the same stamp must not overwrite the evidence
    with pytest.raises(FileExistsError):
        import_local_file(
            input_path=raw,
            source="lianjia",
            dataset="d",
            fetched_at=fetched_at,
            query="q",
            data_dir=tmp_path / "lake",
        )


def test_cli_ingest_file_success_and_catalog(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    raw = tmp_path / "sample.csv"
    raw.write_text("id,price\n1,3200000\n", encoding="utf-8")
    lake = tmp_path / "lake"
    assert (
        cli.main(
            [
                "ingest", "file",
                "--input", str(raw),
                "--source", "SRC-005",
                "--dataset", "chengjiao",
                "--fetched-at", "20260821",
                "--data-dir", str(lake),
            ]
        )
        == 0
    )
    assert "imported 1 rows" in capsys.readouterr().out
    assert cli.main(["catalog", "--data-dir", str(lake)]) == 0
    out = capsys.readouterr().out
    assert "fang_esf/chengjiao@20260821T000000Z" in out


def test_cli_ingest_file_unknown_source_fails(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    raw = tmp_path / "sample.csv"
    raw.write_text("a\n", encoding="utf-8")
    assert (
        cli.main(
            [
                "ingest", "file",
                "--input", str(raw),
                "--source", "nope",
                "--dataset", "d",
                "--fetched-at", "20260821",
                "--data-dir", str(tmp_path / "lake"),
            ]
        )
        == 1
    )
    assert "unknown --source" in capsys.readouterr().out


def test_cli_ingest_file_missing_input_fails(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    assert (
        cli.main(
            [
                "ingest", "file",
                "--input", str(tmp_path / "missing.txt"),
                "--source", "lianjia",
                "--dataset", "d",
                "--fetched-at", "20260821",
                "--data-dir", str(tmp_path / "lake"),
            ]
        )
        == 1
    )
    assert "input not found" in capsys.readouterr().out