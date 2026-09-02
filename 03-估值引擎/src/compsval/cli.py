"""compsval CLI (project skeleton).

Ported and renamed from Philly Fair Measure (fixed SHA e163eba6). Only the
Phase-1 skeleton commands are retained; Philadelphia commands that depend on
deferred or stopped modules (sources, models, diagnostics, opa, api, web,
report, docs_sync) are removed and re-added later as their work packages land.
The dataset registry that powers snapshot-diff/freshness is deferred to the
data-contract work package (WP3); this skeleton ships `system check`, `version`,
`catalog`, `sql`, and `ingest file` (added in WP4-A).

Every command returns a process exit code (0 = success, non-zero = failure).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import subprocess
import sys
from collections import Counter
from collections.abc import Iterator, Sequence
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import pyarrow.parquet as pq

from compsval import __version__, catalog, config
from compsval.contract import registry as contract_registry
from compsval.contract.models import SubjectProperty
from compsval.entities import alias as entities_alias
from compsval.entities import building as entities_building
from compsval.entities import community as entities_community
from compsval.entities import market_series as entities_market_series
from compsval.ingest.binary_snapshot import write_binary_snapshot
from compsval.ingest.floorplan_acceptance import (
    DEFAULT_SAMPLE_SEED,
    DEFAULT_TARGET_SIZE,
    build_acceptance_sample,
    validate_golden_labels,
    write_golden_label_template,
)
from compsval.ingest.floorplan_profile import profile_floorplan
from compsval.ingest.floorplan_selection import (
    PROFILE_ORDINARY_RESIDENTIAL_LATEST,
    build_selection,
)
from compsval.ingest.import_file import import_local_file, resolve_source_dir
from compsval.ingest.manifests import read_derived_manifest
from compsval.ingest.marts_build import build_combined_marts
from compsval.ingest.profile_xlsx import profile_xlsx
from compsval.ingest.stage import MARTS_LAYER, VALID_SALE_FILENAME, data_stage
from compsval.reporting import backtest_report as reporting_backtest_report
from compsval.reporting import markdown as reporting_markdown
from compsval.reporting import run_show as reporting_run_show
from compsval.reporting.envelope import (
    EXIT_OK,
    CommandError,
    InternalCommandError,
    InvalidInputError,
    OutputEnvelope,
    envelope_from_error,
)
from compsval.valuation import (
    aggregation as valuation_aggregation,
)
from compsval.valuation import (
    backtest as valuation_backtest,
)
from compsval.valuation import (
    candidate as valuation_candidate,
)
from compsval.valuation import (
    comparable as valuation_comparable,
)
from compsval.valuation import (
    difference as valuation_difference,
)
from compsval.valuation import (
    estimate as valuation_estimate,
)
from compsval.valuation import (
    review as valuation_review,
)
from compsval.valuation import (
    review_apply as valuation_review_apply,
)
from compsval.valuation import (
    scope as valuation_scope,
)
from compsval.valuation import (
    shadow as valuation_shadow,
)
from compsval.valuation import (
    time_adjustment as valuation_time_adjustment,
)

#: EXTFP4 生产 profile（「冻结支持小区 ∩ 近 12 个月成交」生产子集）
PROFILE_PRODUCTION_WINDOW = "production-supported-window-v1"

#: EXTFP6 全历史生产 profile（「冻结支持小区 ∩ 全部历史成交」，含一致别名匹配）
PROFILE_FULLHISTORY = "supported-community-fullhistory-v1"


def _cmd_version(args: argparse.Namespace) -> int:
    print(f"compsval {__version__}")
    return 0


def _cmd_catalog(args: argparse.Namespace) -> int:
    data_dir = args.data_dir or config.data_dir()
    print("registered example-city sources (data contract):")
    for source in contract_registry.registered_sources():
        print(
            f"  {source.source_id}  {source.name}  "
            f"({source.role.value}/{source.granularity.value}/{source.status.value})"
        )
    print("registered datasets:")
    for dataset in contract_registry.registered_datasets():
        source_ids = ", ".join(dataset.source_ids)
        print(f"  {dataset.dataset}  {dataset.kind}  [{source_ids}]")
    print("evidence snapshots (immutable raw pages):")
    evidence = contract_registry.list_evidence_snapshots()
    if not evidence:
        print("  (no evidence files registered)")
    for snapshot in evidence:
        print(f"  {snapshot.snapshot_id}  {snapshot.format.value}  sha256={snapshot.content_hash}")
    print(f"raw parquet snapshots under {data_dir}:")
    refs = catalog.list_snapshots(data_dir)
    if not refs:
        print(f"  (no raw snapshots under {data_dir})")
    for ref in refs:
        print(f"  {ref.source}/{ref.dataset}@{ref.fetched_at}  {ref.data_path}")
    print("derived parquet tables under {data_dir}:")
    derived = catalog.list_derived(data_dir)
    if not derived:
        print(f"  (no derived tables under {data_dir})")
    for derived_ref in derived:
        print(f"  [{derived_ref.layer}] {derived_ref.view_name}  {derived_ref.path}")
    return 0


def _cmd_sql(args: argparse.Namespace) -> int:
    data_dir = args.data_dir or config.data_dir()
    con = catalog.connect(data_dir)
    try:
        table = con.sql(args.query).pl()
        print(table.head(args.max_rows))
        return 0
    finally:
        con.close()


def _parse_fetched_at(value: str | None, fallback_path: Path) -> datetime:
    """解析 --fetched-at：支持 %Y%m%d 或 %Y%m%dT%H%M%SZ；缺省用文件 mtime。

    时刻格式用于同一数据日内多次导入（如 CX-006-R 补全 provenance 的新快照）。
    """
    if value:
        try:
            return datetime.strptime(value, "%Y%m%d").replace(tzinfo=UTC)
        except ValueError:
            return datetime.strptime(value, "%Y%m%dT%H%M%SZ").replace(tzinfo=UTC)
    return datetime.fromtimestamp(fallback_path.stat().st_mtime, tz=UTC)


def _cmd_ingest_file(args: argparse.Namespace) -> int:
    """Import a local structured file as an immutable raw parquet snapshot.

    Unsupported/unresolvable inputs and stamp collisions fail fast with a clear
    message and a non-zero exit code; the raw evidence file is never modified.
    """
    data_dir = args.data_dir or config.data_dir()
    if not args.input.is_file():
        print(f"[ingest file] input not found: {args.input}")
        return 1
    try:
        source_dir = resolve_source_dir(args.source)
    except Exception as exc:  # noqa: BLE001 - report the user-facing error
        print(f"[ingest file] {exc}")
        return 1
    fetched_at = _parse_fetched_at(args.fetched_at, args.input)
    query = args.query if args.query else f"local file import: {args.input.resolve()}"
    try:
        result = import_local_file(
            input_path=args.input,
            source=source_dir,
            dataset=args.dataset,
            fetched_at=fetched_at,
            query=query,
            data_dir=data_dir,
        )
    except FileExistsError:
        print(
            f"[ingest file] snapshot already exists for {source_dir}/{args.dataset}@"
            f"{fetched_at.isoformat()} (single-snapshot semantics; re-run never "
            f"overwrites evidence)"
        )
        return 1
    print(f"[ingest file] imported {result.manifest.row_count} rows -> {result.directory}")
    return 0


def _xlsx_metadata(path: Path) -> dict[str, object]:
    """XLSX 只读探针：工作表名、成交日期范围、数据行数（技术方案 §5.2 provenance）。

    只读源文件，绝不修改；用于 `compsval ingest binary` 填充原始快照 manifest 的
    provenance 字段（CX-EXTFP1-002 修复）。
    """
    from datetime import date as _date

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet_metadata: list[dict[str, object]] = []
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                sheet_header = next(rows)
            except StopIteration:
                sheet_header = ()
            sheet_data_rows = sum(1 for _ in rows)
            sheet_metadata.append(
                {
                    "sheet": ws.title,
                    "rows": sheet_data_rows,
                    "columns": len(sheet_header),
                    "headers": [str(v) for v in sheet_header],
                }
            )
        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            header = ()
        date_idx = None
        for i, v in enumerate(header):
            if str(v) == "成交日期":
                date_idx = i
                break
        dates: list[_date] = []
        data_rows = 0
        for row in rows:
            data_rows += 1
            if date_idx is not None and date_idx < len(row):
                v = row[date_idx]
                if isinstance(v, datetime):
                    dates.append(v.date())
        return {
            "original_filename": path.name,
            "sheet_names": [s["sheet"] for s in sheet_metadata],
            "sheet_metadata": sheet_metadata,
            "source_row_count": data_rows or None,
            "column_count": len(header) or None,  # 表头列数（§5.2 行列数）
            "data_date_min": min(dates).isoformat() if dates else None,
            "data_date_max": max(dates).isoformat() if dates else None,
        }
    finally:
        wb.close()


def _cmd_ingest_binary(args: argparse.Namespace) -> int:
    """EXTFP1-B：把本地原始字节文件写为不可变二进制快照（data.bin + manifest）。

    原始字节原样复制到数据湖（不入 git），登记 RawSnapshot（format=BINARY，
    mime_type 按扩展名推断或 ``--mime-type`` 显式指定）。XLSX 输入额外做只读
    provenance 探针（工作表名/日期范围/行数）写入 manifest（§5.2）。源文件只读。
    """
    from compsval.ingest.xlsx_parse import PARSE_RULE_VERSION

    data_dir = args.data_dir or config.data_dir()
    if not args.input.is_file():
        print(f"[ingest binary] input not found: {args.input}")
        return 1
    try:
        source_dir = resolve_source_dir(args.source)
    except Exception as exc:  # noqa: BLE001 - report the user-facing error
        print(f"[ingest binary] {exc}")
        return 1
    fetched_at = _parse_fetched_at(args.fetched_at, args.input)
    query = args.query if args.query else f"local binary import: {args.input.resolve()}"
    metadata: dict[str, object] = {}
    if args.input.suffix.lower() == ".xlsx":
        metadata = _xlsx_metadata(args.input)
    # 前序快照 ID（§5.2 provenance）：同一 dataset 早于本次 fetched_at 的最新快照
    prev_snapshot_id: str | None = None
    current_stamp = fetched_at.strftime("%Y%m%dT%H%M%SZ")
    prev_refs = [
        r
        for r in catalog.list_snapshots(data_dir)
        if r.dataset == args.dataset and r.source == source_dir and r.fetched_at < current_stamp
    ]
    if prev_refs:
        latest_prev = max(prev_refs, key=lambda r: r.fetched_at)
        prev_snapshot_id = f"{source_dir}-{args.dataset}-{latest_prev.fetched_at}"
    try:
        result = write_binary_snapshot(
            args.input,
            source=source_dir,
            dataset=args.dataset,
            fetched_at=fetched_at,
            query=query,
            mime_type=args.mime_type,
            root=data_dir,
            original_filename=str(metadata.get("original_filename") or args.input.name),
            sheet_names=metadata.get("sheet_names"),  # type: ignore[arg-type]
            sheet_metadata=metadata.get("sheet_metadata"),  # type: ignore[arg-type]
            source_row_count=(
                int(_src_rows)
                if isinstance(_src_rows := metadata.get("source_row_count"), int)
                else None
            ),
            column_count=(
                int(_col_count)
                if isinstance(_col_count := metadata.get("column_count"), int)
                else None
            ),
            data_date_min=str(metadata["data_date_min"]) if metadata.get("data_date_min") else None,
            data_date_max=str(metadata["data_date_max"]) if metadata.get("data_date_max") else None,
            verification_ref="01-数据/外部数据/画像报告/20260824-外部链家成交画像-EXTFP0-D.json",
            parser_version=PARSE_RULE_VERSION,
            prev_snapshot_id=prev_snapshot_id,
        )
    except FileExistsError:
        print(
            f"[ingest binary] snapshot already exists for {source_dir}/{args.dataset}@"
            f"{fetched_at.isoformat()} (single-snapshot semantics; re-run never "
            f"overwrites evidence)"
        )
        return 1
    r = result.raw_snapshot
    print(
        f"[ingest binary] imported {result.manifest.files[0].size_bytes} bytes "
        f"-> {result.directory}"
    )
    print(
        f"[ingest binary] RawSnapshot {r.snapshot_id} format={r.format.value} "
        f"mime_type={r.mime_type or 'None'} sha256={r.content_hash}"
    )
    if metadata.get("data_date_min"):
        print(
            f"[ingest binary] provenance: sheets={metadata['sheet_names']} "
            f"rows={metadata.get('source_row_count')} "
            f"date=[{metadata['data_date_min']}, {metadata['data_date_max']}]"
        )
    return 0


def _cmd_xlsx_profile(args: argparse.Namespace) -> int:
    """EXTFP0-D：对外部链家成交 Excel 生成只读机器画像报告（JSON）。

    只读输入，绝不修改原始 XLSX；把画像报告写到 ``--out``（原子写入）。stdout
    只打印写入路径；实际画像全部落在 JSON 报告，供复核与下游使用。
    """
    if not args.input.is_file():
        print(f"[xlsx profile] input not found: {args.input}")
        return 1
    try:
        report = profile_xlsx(args.input, out_json=args.out)
    except Exception as exc:  # noqa: BLE001 - report a user-facing failure
        print(f"[xlsx profile] {exc}")
        return 1
    sheet = report.sheets[0]
    print(
        f"[xlsx profile] rows={sheet.data_rows_total} "
        f"普通住宅={sheet.ordinary_residential_count} "
        f"明确排除={sheet.property_use_asserted_excluded} "
        f"用途未知={sheet.property_use_unknown}"
    )
    print(f"[xlsx profile] 画像报告 -> {args.out}")
    return 0


def _cmd_xlsx_floorplan(args: argparse.Namespace) -> int:
    """EXTFP0-E：占位图画像与选择规则冻结（只读，不修改源 XLSX）。

    URL 列表安全解析 + dituFindHouse 占位识别，输出机器占位图画像与冻结的
    选择规则版本到 ``--out``（原子写入）。只读离线，不触发任何下载/OCR。
    """
    if not args.input.is_file():
        print(f"[xlsx floorplan] input not found: {args.input}")
        return 1
    try:
        report = profile_floorplan(args.input, out_json=args.out)
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[xlsx floorplan] {exc}")
        return 1
    o = report.ordinary
    print(
        f"[xlsx floorplan] 普通住宅={o.ordinary_residential_count} "
        f"无URL={o.no_url_count} 解析失败={o.parse_failure_count} "
        f"纯占位={o.placeholder_only_count} 户型图候选={o.floorplan_candidate_count} "
        f"选择规则={report.selection_rule_version}"
    )
    print(f"[xlsx floorplan] 占位图画像与选择规则 -> {args.out}")
    return 0


def _cmd_floorplan_select(args: argparse.Namespace) -> int:
    """EXTFP2-B / EXTFP4：从 staged 普通住宅表生成不可变户型图选择清单（只读，不下载）。

    ``--profile ordinary-residential-latest`` 指向当前 run 的普通住宅子集（或
    ``--run-id`` 覆盖指定 run）。``--profile production-supported-window-v1``
    （EXTFP4）派生「冻结支持小区 ∩ 近 12 个月成交」生产子集。清单即下载契约，
    产出到 ``--out``（默认 data/selection/lianjia_ext/floorplan/selection_manifest.json，
    生产 profile 默认 production_manifest.json，均原子写入）。
    """
    data_dir = args.data_dir or config.data_dir()
    profile = args.profile
    if profile not in (
        PROFILE_ORDINARY_RESIDENTIAL_LATEST,
        PROFILE_PRODUCTION_WINDOW,
        PROFILE_FULLHISTORY,
    ):
        print(
            f"[floorplan select] unknown profile: {profile!r} "
            f"(only {PROFILE_ORDINARY_RESIDENTIAL_LATEST!r} / "
            f"{PROFILE_PRODUCTION_WINDOW!r} / {PROFILE_FULLHISTORY!r} supported)"
        )
        return 1

    from compsval.ingest.xlsx_stage import ORDINARY_FILENAME, read_current_run

    current = read_current_run(data_dir)
    if current is None:
        print(
            "[floorplan select] no current run pointer "
            "(staged/lianjia_ext/current.json); run compsval xlsx stage first"
        )
        return 1

    if args.run_id:
        run_id = args.run_id
        rel = f"runs/run_{run_id}/{ORDINARY_FILENAME}"
    else:
        run_id = current.get("run_id")
        rel = current.get("ordinary_residential") or ""
    parquet_path = data_dir / "staged" / "lianjia_ext" / rel
    if not parquet_path.is_file():
        print(f"[floorplan select] staged parquet not found: {parquet_path}")
        return 1

    if profile == PROFILE_PRODUCTION_WINDOW:
        return _floorplan_select_production(data_dir, parquet_path, args.out)

    if profile == PROFILE_FULLHISTORY:
        return _floorplan_select_fullhistory(data_dir, parquet_path, args.out)

    out = args.out or (
        data_dir / "selection" / "lianjia_ext" / "floorplan" / "selection_manifest.json"
    )
    try:
        manifest = build_selection(parquet_path, run_id=run_id, out_json=out)
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan select] {exc}")
        return 1
    print(
        f"[floorplan select] run={manifest.run_id} 记录={manifest.record_count} "
        f"资产={manifest.asset_count} 违规域名违法URL={manifest.forbidden_domain_count} "
        f"hash={manifest.record_ids_hash[:12]}… 选择规则={manifest.selection_rule_version}"
    )
    print(
        f"[floorplan select] 成交{manifest.date_range_min}~{manifest.date_range_max} "
        f"预计下载≈{manifest.estimated_download_bytes / 1048576:.1f}MiB "
        f"存储上限={manifest.storage_cap_bytes / 1048576:.1f}MiB "
        f"成本上限=¥{manifest.budget_cap_yuan:.2f}"
    )
    print(f"[floorplan select] 选择清单 -> {out}")
    return 0


def _floorplan_select_production(data_dir: Path, parquet_path: Path, out_arg: Path | None) -> int:
    """EXTFP4 生产 profile：派生「冻结支持小区 ∩ 近 12 个月」子集清单（只读，不下载）。

    输出 production_manifest.json + SHA256 旁证 + production_exclusion_report.json；
    派生锚点与既有全量 selection_manifest.json 核对（不一致即拒绝派生）。
    """
    from compsval.ingest.floorplan_production import build_production_selection

    floorplan_dir = data_dir / "selection" / "lianjia_ext" / "floorplan"
    out = out_arg or floorplan_dir / "production_manifest.json"
    exclusion_out = out.with_name("production_exclusion_report.json")
    full_manifest_path = floorplan_dir / "selection_manifest.json"
    try:
        production, exclusion = build_production_selection(
            parquet_path,
            data_dir / "entities",
            out_json=out,
            exclusion_out_json=exclusion_out,
            full_manifest_path=full_manifest_path if full_manifest_path.is_file() else None,
        )
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan select] {exc}")
        return 1
    print(
        f"[floorplan select] 生产清单 run={production.run_id} "
        f"基线={production.baseline_record_count} 记录={production.record_count} "
        f"资产={production.asset_count} hash={production.record_ids_hash[:12]}… "
        f"规则={production.selection_rule_version}"
    )
    print(
        f"[floorplan select] 小区命中={production.matched_community_counts} "
        f"预算外推=¥{production.budget_expected_yuan:.4f} "
        f"清单上限=¥{production.budget_cap_yuan:.2f} attempt上限={production.attempt_cap}"
    )
    print(f"[floorplan select] 排除清单 -> {exclusion_out}")
    print(f"[floorplan select] 选择清单 -> {out}")
    _ = exclusion
    return 0


def _floorplan_select_fullhistory(data_dir: Path, parquet_path: Path, out_arg: Path | None) -> int:
    """EXTFP6 全历史 profile：派生「冻结支持小区 ∩ 全部历史成交」子集清单（只读，不下载）。

    匹配含 community_alias 中 conflict_status=一致 的别名（示例小区130A区/B区）；不限
    成交时间窗；change 级预算硬上限 ≤¥10（fail-closed）。输出
    production_manifest_fullhistory.json + SHA256 旁证 + 对应排除清单报告；
    派生锚点与既有全量 selection_manifest.json 核对（不一致即拒绝派生）。
    """
    from compsval.ingest.floorplan_production import (
        EXTFP6_CHANGE_BUDGET_CAP_YUAN,
        EXTFP6_CHANGE_REF,
        FULLHISTORY_FILTER_TEXT,
        FULLHISTORY_PROFILE,
        FULLHISTORY_RULE_TEXT,
        FULLHISTORY_RULE_VERSION,
        build_production_selection,
    )

    floorplan_dir = data_dir / "selection" / "lianjia_ext" / "floorplan"
    out = out_arg or floorplan_dir / "production_manifest_fullhistory.json"
    exclusion_out = out.with_name("production_exclusion_report_fullhistory.json")
    full_manifest_path = floorplan_dir / "selection_manifest.json"
    try:
        production, exclusion = build_production_selection(
            parquet_path,
            data_dir / "entities",
            out_json=out,
            exclusion_out_json=exclusion_out,
            full_manifest_path=full_manifest_path if full_manifest_path.is_file() else None,
            date_min=None,
            date_max=None,
            profile=FULLHISTORY_PROFILE,
            change_ref=EXTFP6_CHANGE_REF,
            workpackage_ref="EXTFP6",
            change_budget_cap_yuan=EXTFP6_CHANGE_BUDGET_CAP_YUAN,
            selection_rule_version=FULLHISTORY_RULE_VERSION,
            selection_rule_text=FULLHISTORY_RULE_TEXT,
            filter_condition_text=FULLHISTORY_FILTER_TEXT,
        )
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan select] {exc}")
        return 1
    print(
        f"[floorplan select] 全历史生产清单 run={production.run_id} "
        f"基线={production.baseline_record_count} 记录={production.record_count} "
        f"资产={production.asset_count} hash={production.record_ids_hash[:12]}… "
        f"规则={production.selection_rule_version}"
    )
    print(
        f"[floorplan select] 小区命中={production.matched_community_counts} "
        f"预算外推=¥{production.budget_expected_yuan:.4f} "
        f"清单上限=¥{production.budget_cap_yuan:.2f} attempt上限={production.attempt_cap} "
        f"change硬上限=¥{production.change_budget_cap_yuan:.2f}"
    )
    print(f"[floorplan select] 排除清单 -> {exclusion_out}")
    print(f"[floorplan select] 选择清单 -> {out}")
    _ = exclusion
    return 0


def _cmd_floorplan_acceptance(args: argparse.Namespace) -> int:
    """EXTFP3-G：从 staged 普通住宅表生成 300 张确定性分层验收抽样清单（只读，不下载）。

    ``--target`` 默认 300、``--seed`` 默认 20260825（与 EXTFP3-G 冻结锚点一致）。
    抽样结果是与 SelectionManifest 兼容的清单，可被 ``compsval floorplan download`` 消费；
    ``--golden-csv`` 指定时顺带生成人工标注黄金标签 CSV 模板。
    """
    data_dir = args.data_dir or config.data_dir()
    profile = args.profile
    if profile != PROFILE_ORDINARY_RESIDENTIAL_LATEST:
        print(
            f"[floorplan acceptance] unknown profile: {profile!r} "
            f"(only {PROFILE_ORDINARY_RESIDENTIAL_LATEST!r} supported)"
        )
        return 1

    from compsval.ingest.xlsx_stage import ORDINARY_FILENAME, read_current_run

    current = read_current_run(data_dir)
    if current is None:
        print(
            "[floorplan acceptance] no current run pointer "
            "(staged/lianjia_ext/current.json); run compsval xlsx stage first"
        )
        return 1

    if args.run_id:
        run_id = args.run_id
        rel = f"runs/run_{run_id}/{ORDINARY_FILENAME}"
    else:
        run_id = current.get("run_id")
        rel = current.get("ordinary_residential") or ""
    parquet_path = data_dir / "staged" / "lianjia_ext" / rel
    if not parquet_path.is_file():
        print(f"[floorplan acceptance] staged parquet not found: {parquet_path}")
        return 1

    source_manifest = (
        data_dir / "selection" / "lianjia_ext" / "floorplan" / "selection_manifest.json"
    )
    out = args.out or (
        data_dir / "selection" / "lianjia_ext" / "floorplan" / "acceptance_manifest.json"
    )
    try:
        manifest = build_acceptance_sample(
            parquet_path,
            target=args.target,
            seed=args.seed,
            out_json=out,
            source_manifest_path=source_manifest if source_manifest.is_file() else None,
        )
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan acceptance] {exc}")
        return 1

    if args.golden_csv:
        try:
            write_golden_label_template(manifest, args.golden_csv)
        except Exception as exc:  # noqa: BLE001 - user-facing failure
            print(f"[floorplan acceptance] 黄金标签模板生成失败: {exc}")
            return 1

    if args.golden_validate:
        try:
            validation = validate_golden_labels(
                args.golden_validate,
                manifest,
                out_json=out.with_name("golden_validation.json"),
            )
        except Exception as exc:  # noqa: BLE001 - user-facing failure
            print(f"[floorplan acceptance] 黄金标签校验失败: {exc}")
            return 1
        print(
            f"[floorplan acceptance] 黄金标签校验 valid={validation.valid} "
            f"rows={validation.rows_ok}/{validation.expected_samples} "
            f"房间={validation.room_count_total} 有面积={validation.area_present_count} "
            f"缺={len(validation.missing_samples)} 非法={len(validation.invalid_entries)}"
        )
        if not validation.valid:
            return 2

    print(
        f"[floorplan acceptance] run={manifest.run_id} 记录={manifest.record_count} "
        f"资产={manifest.asset_count} 目标={manifest.target_size} 种子={manifest.random_seed} "
        f"hash={manifest.record_ids_hash[:12]}… 规则={manifest.sampling_rule_version}"
    )
    print(
        f"[floorplan acceptance] 成交{manifest.date_range_min}~{manifest.date_range_max} "
        f"裁剪额外资产={manifest.trimmed_extra_assets} "
        f"预计下载≈{manifest.estimated_download_bytes / 1048576:.1f}MiB "
        f"成本上限=¥{manifest.budget_cap_yuan:.2f}"
    )
    print(f"[floorplan acceptance] 抽样清单 -> {out}")
    if args.golden_csv:
        print(f"[floorplan acceptance] 黄金标签模板 -> {args.golden_csv}")
    return 0


def _cmd_floorplan_download(args: argparse.Namespace) -> int:
    """EXTFP2-C：户型图下载器与状态机（离线/在线按注入 transport 决定）。

    从 ``--selection``(selection_manifest JSON) 读取待下载资产清单，按域名白名单、
    幂等键、有界并发、指数退避重试与断点续跑把原始字节落盘到 ``--out``（默认
    data/download/lianjia_ext/floorplan/），并维护 download_state.json。
    返回码：全成功 0；部分失败仍 0（发出摘要）；缺清单文件 1；参数非法 2。

    EXTFP2-C 只交付下载器 + 状态机 + 离线 mock 测试，不执行任何真实网络下载
    （真实 10 张试跑属 EXTFP2-E）。
    """
    selection = args.selection
    if not selection.is_file():
        print(f"[floorplan download] selection manifest not found: {selection}")
        return 1
    for flag, value in (("--max-concurrency", args.max_concurrency), ("--retries", args.retries)):
        if value < 1:
            print(f"[floorplan download] invalid {flag}={value}（必须 >=1）")
            return 2
    if args.timeout <= 0:
        print(f"[floorplan download] invalid --timeout={args.timeout}（必须 >0）")
        return 2

    from compsval.ingest.floorplan_download import (
        DOWNLOADER_VERSION,
        DownloadState,
        run_download,
    )
    from compsval.ingest.floorplan_selection import SelectionManifest

    try:
        manifest = SelectionManifest.model_validate(
            json.loads(selection.read_text(encoding="utf-8"))
        )
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan download] invalid selection manifest: {exc}")
        return 1

    data_dir = args.data_dir or config.data_dir()
    out = args.out or (data_dir / "download" / "lianjia_ext" / "floorplan")
    try:
        record = run_download(
            manifest,
            out,
            max_concurrency=args.max_concurrency,
            max_attempts=args.retries,
            timeout=args.timeout,
            force_new_run=bool(args.force_new_run),
            manifest_path=selection,
        )
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan download] {exc}")
        return 1

    ok = record.state_counts.get(DownloadState.DOWNLOADED.value, 0)
    failed = record.state_counts.get(DownloadState.DOWNLOAD_FAILED.value, 0)
    print(
        f"[floorplan download] run={record.run_id} 资产={len(record.tasks)} "
        f"成功下载={ok} 失败={failed} 版本={DOWNLOADER_VERSION}"
    )
    print(f"[floorplan download] 状态与任务清单 -> {record.run_dir}")
    return 0


def _cmd_floorplan_asset(args: argparse.Namespace) -> int:
    """EXTFP2-D：原图资产化与字节校验（魔数/MIME/尺寸、扩展名由字节决定、SHA256 复核）。

    读取 ``--run`` 目录（compsval floorplan download 的输出：download_state.json + *.img），
    校验并落盘原图到 ``--out``（数据湖 ``raw/source=lianjia_ext/dataset=floorplan_image/
    batch_id=<id>/``），生成不可变 ``floorplan_asset_manifest.json``。
    返回码：成功 0；缺下载运行 1；异常 2。
    """
    run_dir = args.run
    if not run_dir.is_dir():
        print(f"[floorplan asset] download run dir not found: {run_dir}")
        return 1
    if (
        not (run_dir / "download_state.json").is_file()
        and not (run_dir / "download_run.json").is_file()
    ):
        print(f"[floorplan asset] no download state in {run_dir}")
        return 1

    from compsval.ingest.floorplan_asset import (
        ASSET_MANIFEST_FILENAME,
        ASSET_STAGED_FILENAME,
        AssetStatus,
        build_asset_manifest,
        write_staged_asset_table,
    )

    data_dir = args.data_dir or config.data_dir()
    out = args.out or (data_dir / "raw" / "source=lianjia_ext" / "dataset=floorplan_image")
    try:
        run = build_asset_manifest(run_dir, out, batch_id=args.batch_id or None)
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan asset] {exc}")
        return 2

    staged_path = write_staged_asset_table(run, data_dir)

    valid = run.counts.get(AssetStatus.DOWNLOADED.value, 0)
    invalid = run.counts.get(AssetStatus.IMAGE_INVALID.value, 0)
    notavail = run.counts.get(AssetStatus.NOT_AVAILABLE.value, 0)
    batch_dir = out / f"batch_id={run.batch_id}"
    print(
        f"[floorplan asset] batch={run.batch_id} 有效={valid} 无效={invalid} "
        f"未下载={notavail} 版本={run.rules_version}"
    )
    print(f"[floorplan asset] 资产 manifest -> {batch_dir / ASSET_MANIFEST_FILENAME}")
    print(f"[floorplan asset] staged 资产表 -> {staged_path}（{ASSET_STAGED_FILENAME}）")
    return 0


def _cmd_floorplan_e2e(args: argparse.Namespace) -> int:
    """EXTFP2-E：样本只读登记 + 10 张真实试跑完整链路。

    串起「样本只读登记 → 从全量 selection_manifest 重建 10 张子集 → 真实下载 →
    原图资产与校验 → 下载质量报告 / 链路证据」，并写入 ``e2e_bundle.json`` 聚合证据。
    仅本 CL 触发真实网络下载（子集 10 张 + 域名白名单 fail-closed）。
    返回码：成功 0；缺输入 1；参数非法/链路异常 2。
    """
    selection = args.selection
    full_manifest_path = selection
    if not full_manifest_path.is_file():
        print(f"[floorplan e2e] selection manifest not found: {full_manifest_path}")
        return 1
    if not args.sample_dir.is_dir():
        print(f"[floorplan e2e] 样本目录 not found: {args.sample_dir}")
        return 1
    sample_list = args.sample_list or (args.sample_dir / "样本来源清单.md")
    if not sample_list.is_file():
        print(f"[floorplan e2e] 样本来源清单 not found: {sample_list}")
        return 1
    for flag, value in (
        ("--max-concurrency", args.max_concurrency),
        ("--retries", args.retries),
    ):
        if value < 1:
            print(f"[floorplan e2e] invalid {flag}={value}（必须 >=1）")
            return 2
    if args.timeout <= 0:
        print(f"[floorplan e2e] invalid --timeout={args.timeout}（必须 >0）")
        return 2
    if args.expected_count < 1:
        print(f"[floorplan e2e] invalid --expected-count={args.expected_count}（必须 >=1）")
        return 2

    from datetime import UTC, datetime

    from compsval.ingest.floorplan_asset import (
        build_asset_manifest,
        write_cumulative_staged_asset_table,
    )
    from compsval.ingest.floorplan_download import (
        DOWNLOADER_VERSION,
        DownloadState,
        run_download,
    )
    from compsval.ingest.floorplan_e2e import (
        SUBSET_FILENAME,
        E2eBundle,
        aggregate_bundle_counts,
        build_download_manifest,
        build_download_quality_report,
        build_subset_manifest,
        collect_existing_downloaded_asset_ids,
        load_asset_run_manifests,
        parse_sample_list,
        register_samples,
        resolve_missing_assets,
    )
    from compsval.ingest.floorplan_selection import (
        DOMAIN_WHITELIST,
        SelectionManifest,
    )

    try:
        full_manifest = SelectionManifest.model_validate(
            json.loads(full_manifest_path.read_text(encoding="utf-8"))
        )
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan e2e] invalid selection manifest: {exc}")
        return 1

    data_dir = args.data_dir or config.data_dir()
    out_dir = args.out or (data_dir / "selection" / "lianjia_ext" / "floorplan" / "e2e")
    out_dir.mkdir(parents=True, exist_ok=True)

    # 1) 样本只读登记（只读，不联网）
    sample_files = parse_sample_list(args.sample_dir, sample_list)
    registration = register_samples(full_manifest, sample_files)
    reg_path = out_dir / "sample_registration.json"
    reg_path.write_text(
        registration.model_dump_json(indent=2, ensure_ascii=False), encoding="utf-8"
    )

    # 2) 重建 expected_count 张子集（seed 命中不足按稳定顺序补选，仍不足 fail-closed 拒绝）
    expected_count = args.expected_count
    subset_path = out_dir / SUBSET_FILENAME
    try:
        subset = build_subset_manifest(
            full_manifest,
            sample_files,
            subset_path,
            expected_count=expected_count,
        )
    except ValueError as exc:  # 补选后仍不足 expected_count → fail-closed：禁止成功子集，不触网
        print(f"[floorplan e2e] {exc}")
        return 2

    # 3) 真实下载：仅补下子集尚缺失的资产（既有成功资产绝不重复下发，网络预算 <=expected_count）
    raw_dir = data_dir / "raw" / "source=lianjia_ext" / "dataset=floorplan_image"
    dl_dir = out_dir / "download"
    existing_ids = collect_existing_downloaded_asset_ids(raw_dir)
    missing = resolve_missing_assets(
        subset.records,
        full_manifest.snapshot_ref,
        existing_ids,
    )
    new_download_run_id: str | None = None
    new_asset_batch_id: str | None = None
    if missing:
        print(
            f"[floorplan e2e] 子集={subset.asset_count} 已存在={len(existing_ids)} "
            f"本次补下缺失={len(missing)}（累计不超过 {expected_count}）"
        )
        dl_manifest = build_download_manifest(full_manifest, missing)
        dl_record = run_download(
            dl_manifest,
            dl_dir,
            max_concurrency=args.max_concurrency,
            max_attempts=args.retries,
            timeout=args.timeout,
            force_new_run=True,  # 新不可变 run/batch，绝不覆盖既有 5 张证据
            manifest_path=subset_path,
        )
        new_download_run_id = dl_record.run_id
        dl_run_dir = Path(dl_record.run_dir)
        # 4) 校验新下载批次并落新 batch（原始字节 + 不可变 asset manifest）
        new_asset_run = build_asset_manifest(dl_run_dir, raw_dir, batch_id=args.batch_id or None)
        new_asset_batch_id = new_asset_run.batch_id
    else:
        print("[floorplan e2e] 子集资产均已落盘，无需网络下载（本次仅刷新累计证据）")

    # 5) 累计证据：跨全部 batch_id 聚合计数 + staged 累计表 + 质量报告 + e2e_bundle
    manifest_runs = load_asset_run_manifests(raw_dir)
    if not manifest_runs:
        print("[floorplan e2e] 无任何资产 manifest，无法生成累计证据")
        return 2
    download_counts, asset_counts = aggregate_bundle_counts(manifest_runs)
    aggregated_run_ids = sorted({str(m.get("download_run_id")) for m in manifest_runs})
    aggregated_batch_ids = sorted({str(m.get("batch_id")) for m in manifest_runs})

    # 累计 staged 表（每行保留各自 batch_id 血缘；原子切换 current 指针）
    staged_path = write_cumulative_staged_asset_table(manifest_runs, data_dir)

    created_at = datetime.now(UTC).isoformat()
    bundle = E2eBundle(
        created_at=created_at,
        selection_ref=str(full_manifest_path),
        selection_rule_version=full_manifest.selection_rule_version,
        subset_path=str(subset_path),
        subset_asset_count=subset.asset_count,
        expected_count=expected_count,
        download_run_id=new_download_run_id or aggregated_run_ids[-1],
        downloader_version=DOWNLOADER_VERSION,
        download_state_counts=download_counts,
        asset_batch_id=new_asset_batch_id or aggregated_batch_ids[-1],
        asset_rules_version=manifest_runs[-1].get("rules_version", ""),
        asset_counts=asset_counts,
        domain_whitelist=sorted(DOMAIN_WHITELIST),
        aggregated_download_run_ids=aggregated_run_ids,
        aggregated_batch_ids=aggregated_batch_ids,
    )
    qrep_json, qrep_md = build_download_quality_report(
        bundle,
        sample_registration=registration,
        asset_manifests=manifest_runs,
        local_files=sample_files,
        out_dir=out_dir,
    )
    bundle_path = out_dir / "e2e_bundle.json"
    bundle_path.write_text(bundle.model_dump_json(indent=2, ensure_ascii=False), encoding="utf-8")

    ok = download_counts.get(DownloadState.DOWNLOADED.value, 0)
    failed = 0
    valid = asset_counts.get("DOWNLOADED", 0)
    print(
        f"[floorplan e2e] subset={subset.asset_count}/{expected_count} 累计下载成功={ok} "
        f"失败={failed} 资产有效={valid} 样本SHA256比对见质量报告"
    )
    print(f"[floorplan e2e] 子集清单 -> {subset_path}")
    print(f"[floorplan e2e] 聚合 run_id -> {', '.join(aggregated_run_ids)}")
    print(f"[floorplan e2e] 聚合 batch_id -> {', '.join(aggregated_batch_ids)}")
    print(f"[floorplan e2e] 质量报告 -> {qrep_json} / {qrep_md}")
    print(f"[floorplan e2e] 链路证据 -> {bundle_path}")
    print(f"[floorplan e2e] staged 资产表 -> {staged_path}")
    return 0


def _cmd_floorplan_ocr(args: argparse.Namespace) -> int:
    """EXTFP3-B：Qwen OCR 请求器与运行记录（技术方案 §8.3/§9.2/§11.3/§12/§15）。

    从 ``--asset-manifest``（floorplan_asset_manifest.json）读取已资产化的原图，
    按冻结合同调用 Qwen OCR（网络=是、付费=是，需工作包授权），逐张保存原始响应与
    ``floorplan_ocr_run`` 运行记录；幂等键断点续跑 + 成本门禁 fail-closed。
    返回码：成功 0；缺 manifest / 密钥 1；参数非法 2。
    """
    manifest_path = args.asset_manifest
    if not manifest_path.is_file():
        print(f"[floorplan ocr] asset manifest not found: {manifest_path}")
        return 1
    if args.retries < 1:
        print(f"[floorplan ocr] invalid --retries={args.retries}（必须 >=1）")
        return 2
    if args.timeout <= 0:
        print(f"[floorplan ocr] invalid --timeout={args.timeout}（必须 >0）")
        return 2

    from compsval.ingest.floorplan_ocr import (
        OCR_REQUESTER_VERSION,
        run_ocr_batch,
    )
    from compsval.ingest.floorplan_ocr_contract import (
        load_ocr_run_config,
        read_dashscope_api_key,
    )

    # §12：OCR 命令显式标注会访问网络/付费；启动前检查密钥与配置
    print("[floorplan ocr] 本命令会调用真实 Qwen OCR（网络=是、付费=是，需工作包授权）")
    try:
        read_dashscope_api_key()
    except KeyError as exc:
        print(f"[floorplan ocr] {exc}")
        return 1
    config_obj = load_ocr_run_config(args.config) if args.config else None

    data_dir = args.data_dir or config.data_dir()
    out = args.out or (data_dir / "raw" / "source=lianjia_ext" / "dataset=floorplan_ocr_run")
    try:
        record = run_ocr_batch(
            manifest_path,
            out,
            config=config_obj,
            force_new_run=bool(args.force_new_run),
            timeout=args.timeout,
            max_attempts=args.retries,
            concurrency=args.concurrency,
        )
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan ocr] {exc}")
        return 1

    # §12：标准输出只写 JSON；进度日志已走 stderr/摘要
    print(
        json.dumps(
            {
                "ocr_run_id": record.ocr_run_id,
                "requester_version": OCR_REQUESTER_VERSION,
                "asset_count": len(record.tasks),
                "state_counts": record.state_counts,
                "cost": record.cost,
                "concurrency": args.concurrency,
                "performance": record.performance,
                "run_dir": record.run_dir,
            },
            ensure_ascii=False,
            sort_keys=True,
        )
    )
    print(f"[floorplan ocr] 运行记录 -> {record.run_dir}", file=sys.stderr)
    return 0


def _cmd_floorplan_transcribe(args: argparse.Namespace) -> int:
    """EXTFP3-D：确定性转录（逐词表 → 房间标注表 + 参与字段回填）。

    读取 OCR 运行的 ``floorplan_ocr_word.parquet``，用确定性转录解析器（纯函数）生成
    房间标注表，原子写 ``staged/floorplan_room_annotation.parquet`` 并回填逐词表
    ``participates_in_field``（只追加派生字段，不覆盖 C/D 元数据）。不触网、不付费。
    返回码：成功 0；缺运行/词表 1；参数非法 2。
    """
    if not args.run.is_dir():
        print(f"[floorplan transcribe] run dir not found: {args.run}", file=sys.stderr)
        return 1
    data_dir = args.data_dir or config.data_dir()
    word_table = args.word_table
    if word_table is None:
        candidates = [
            args.run / "floorplan_ocr_word.parquet",
            data_dir / "staged" / "floorplan_ocr_word.parquet",
        ]
        word_table = next((c for c in candidates if c.is_file()), None)
    if word_table is None:
        print(
            "[floorplan transcribe] 未找到逐词表 floorplan_ocr_word.parquet"
            "（可 --word-table 显式指定）",
            file=sys.stderr,
        )
        return 1
    try:
        from compsval.ingest.floorplan_transcribe import (
            TRANSCRIBE_PARSER_VERSION,
            transcribe_word_table,
        )

        stats = transcribe_word_table(word_table, data_dir=data_dir)
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan transcribe] {exc}", file=sys.stderr)
        return 1
    payload = stats.model_dump(mode="json")
    payload["parser_version"] = TRANSCRIBE_PARSER_VERSION
    print(json.dumps(payload, ensure_ascii=False, sort_keys=True))
    print(
        f"[floorplan transcribe] 标注表 -> {stats.table_path or '(未落盘)'}；"
        f"参与字段回填 -> {stats.word_participation_path or '(未落盘)'}",
        file=sys.stderr,
    )
    return 0


def _cmd_floorplan_verify(args: argparse.Namespace) -> int:
    """EXTFP3-E：自动一致性检查 + 质量报告（§9.5/§14）。

    对一次 OCR 运行执行全部一致性检查并生成质量报告 JSON；可选回填标注表
    ``consistency_status``（只追加派生字段，不覆盖原始标注）。不触网、不付费。
    返回码：成功 0；缺运行 1；参数非法 2。
    """
    if not args.run.is_dir():
        print(f"[floorplan verify] run dir not found: {args.run}", file=sys.stderr)
        return 1
    data_dir = args.data_dir or config.data_dir()
    repeat_annotations = None
    if args.repeat_annotations is not None:
        if not args.repeat_annotations.is_file():
            print(
                f"[floorplan verify] repeat annotations not found: {args.repeat_annotations}",
                file=sys.stderr,
            )
            return 1
        from compsval.ingest.floorplan_verify import read_annotation_table

        repeat_annotations = read_annotation_table(args.repeat_annotations)
    try:
        from compsval.ingest.floorplan_verify import verify_run

        report = verify_run(
            args.run,
            data_dir=data_dir,
            word_table_path=args.word_table,
            annotation_table_path=args.annotation_table,
            asset_manifest_path=args.asset_manifest,
            staged_table_path=args.staged_table,
            repeat_annotations=repeat_annotations,
            write_consistency=bool(args.write_consistency),
        )
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan verify] {exc}", file=sys.stderr)
        return 1
    print(report.model_dump_json(indent=2, ensure_ascii=False))
    check_summary = ", ".join(f"{c.check_id}={c.status}" for c in report.checks)
    print(f"[floorplan verify] overall={report.overall}；{check_summary}", file=sys.stderr)
    return 0


def _cmd_floorplan_status(args: argparse.Namespace) -> int:
    """EXTFP3-E：状态机聚合与失败分类（§12）。

    读取 OCR 运行记录，按状态机聚合状态计数并显式分类失败（可重试/可追溯/可分类），
    供调试与断点续跑。不触网、不付费。返回码：成功 0；缺运行 1。
    """
    if not args.run.is_dir():
        print(f"[floorplan status] run dir not found: {args.run}", file=sys.stderr)
        return 1
    try:
        from compsval.ingest.floorplan_ocr_parse import load_ocr_run_record
        from compsval.ingest.floorplan_verify import build_status_report

        run = load_ocr_run_record(args.run)
        report = build_status_report(run, run_dir=args.run)
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan status] {exc}", file=sys.stderr)
        return 1
    print(report.model_dump_json(indent=2, ensure_ascii=False))
    print(
        f"[floorplan status] run={report.ocr_run_id} 状态={report.state_counts} "
        f"失败={len(report.failures)} 可重试={len(report.retryable_tasks)}",
        file=sys.stderr,
    )
    return 0


def _cmd_floorplan_freeze(args: argparse.Namespace) -> int:
    """EXTFP5：数据冻结登记（只读盘点 → 生成 freeze manifest → 只读校验 → 写版本指针）。

    全部离线、零网络零付费；不对任何既有产物执行写操作，只新建
    ``data/versions/`` 下的 manifest / 指针 / 报告。校验或披露缺项失败时
    返回非 0 且不写指针（版本保持未发布）。
    """
    from compsval.ingest.data_freeze import (
        FreezeManifest,
        load_manifest,
        run_freeze,
        verify_manifest,
    )

    data_dir = args.data_dir or config.data_dir()
    try:
        if args.verify_only is not None:
            from compsval.ingest.data_freeze import _repo_root_of

            manifest = load_manifest(args.verify_only)
            verification = verify_manifest(manifest, _repo_root_of(data_dir))
            print(verification.model_dump_json(indent=2, ensure_ascii=False))
            if not verification.ok:
                print("[floorplan freeze] verify FAILED（版本未发布）", file=sys.stderr)
                return 1
            print("[floorplan freeze] verify PASS（只读校验一致）", file=sys.stderr)
            return 0
        result = run_freeze(data_dir)
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[floorplan freeze] {exc}", file=sys.stderr)
        return 1
    if not result["ok"]:
        for gap in result["gaps"]:
            print(f"[floorplan freeze] {gap}", file=sys.stderr)
        print(
            f"[floorplan freeze] 冻结失败（stage={result['stage']}，版本保持未发布）",
            file=sys.stderr,
        )
        return 1
    print(
        json.dumps(
            {
                "ok": True,
                "version_id": FreezeManifest.model_validate_json(
                    result["manifest"].read_text(encoding="utf-8")
                ).version_id,
                "manifest": str(result["manifest"]),
                "sidecar": str(result["sidecar"]),
                "pointer": str(result["pointer"]),
                "report": str(result["report"]),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


def _cmd_floorplan_unrecognized(args: argparse.Namespace) -> int:
    """``compsval floorplan`` 未带子命令时的兜底：给用法而非崩溃（RV-EXTFP2-C-01#F1）。"""
    print(
        "usage: compsval floorplan {select,download,asset,e2e,ocr,transcribe,verify,status,freeze} "
        "[options]",
        file=sys.stderr,
    )
    return 2


def _cmd_xlsx_unrecognized(args: argparse.Namespace) -> int:
    """``compsval xlsx`` 未带子命令时的兜底：给用法而非崩溃（RV-EXTFP0-D-01#F1）。"""
    print(
        "usage: compsval xlsx {profile,floorplan,parse,stage,quality} --input <XLSX> --out <JSON>",
        file=sys.stderr,
    )
    return 2


def _cmd_xlsx_parse(args: argparse.Namespace) -> int:
    """EXTFP1-C：全量逐行解析（§6 字段映射 + 缺失语义 + 户型图 URL 衔接）。

    流式解析原始 XLSX（只读，不修改源文件），输出机器摘要 JSON：行数守恒、
    用途分布、字段缺失/解析失败统计、映射规则版本、源 SHA256 血缘与可选抽样
    记录。全量记录不落盘——staged 表由 EXTFP1-D 承接。
    """
    from compsval.ingest.xlsx_parse import (
        iter_parse_xlsx,
        summarize,
    )

    if not args.input.is_file():
        print(f"[xlsx parse] input not found: {args.input}")
        return 1
    sample_size = args.sample or 0
    sample: list[dict] = []
    sha = hashlib.sha256()
    try:
        with args.input.open("rb") as fh:
            for chunk in iter(lambda: fh.read(1 << 20), b""):
                sha.update(chunk)
        records = iter_parse_xlsx(args.input)
        if sample_size:

            def _sample_wrap(inner: Iterator[Any] = records) -> Iterator[Any]:
                for rec in inner:
                    if len(sample) < sample_size:
                        sample.append(rec.model_dump(mode="json"))
                    yield rec

            records = _sample_wrap()
        summary = summarize(records, source_sha256=sha.hexdigest())
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[xlsx parse] {exc}")
        return 1
    report = {
        "workpackage": "EXTFP1-C",
        "name": "全量逐行解析摘要",
        "parse_rule_version": summary.parse_rule_version,
        "source_path": str(args.input.resolve()),
        "source_sha256": summary.source_sha256,
        "data_rows_total": summary.data_rows_total,
        "parsed_count": summary.parsed_count,
        "ordinary_residential_count": summary.ordinary_residential_count,
        "property_use_distribution": summary.property_use_distribution,
        "field_status_counts": summary.field_status_counts,
        "sample_records": sample,
    }
    args.out.parent.mkdir(parents=True, exist_ok=True)
    work = args.out.with_name(args.out.name + ".incomplete")
    work.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    work.replace(args.out)
    print(
        f"[xlsx parse] rows={summary.data_rows_total} "
        f"普通住宅={summary.ordinary_residential_count} 抽样={len(sample)} "
        f"规则={summary.parse_rule_version}"
    )
    print(f"[xlsx parse] 解析摘要 -> {args.out}")
    return 0


def _cmd_xlsx_stage(args: argparse.Namespace) -> int:
    """EXTFP1-D：全量解析落 staged 表（sale_record + ordinary_residential + 血缘）。

    只读源 XLSX，流式全量解析后按不可变 ``run_id`` 写入
    ``data_dir/staged/lianjia_ext/runs/run_<id>/`` 两张 parquet 表及各自
    DerivedManifest（结构化 inputs 指向实际二进制快照 + 解析规则版本），完成后
    原子切换 current 指针。旧 run 产物永久保留，绝不覆盖（技术方案 §16）。
    """
    from compsval.ingest.manifests import InputRef
    from compsval.ingest.xlsx_stage import (
        ORDINARY_FILENAME,
        SALE_RECORD_FILENAME,
        stage_xlsx,
    )

    data_dir = args.data_dir or config.data_dir()
    if not args.input.is_file():
        print(f"[xlsx stage] input not found: {args.input}")
        return 1
    # 结构化血缘：指向最近的 lianjia_ext/chengjiao_xlsx 二进制快照（含内容 hash）
    inputs: list[InputRef] = []
    refs = [
        r
        for r in catalog.list_snapshots(data_dir)
        if r.dataset == "chengjiao_xlsx" and r.source == "lianjia_ext"
    ]
    if refs:
        latest = max(refs, key=lambda r: r.fetched_at)
        manifest = latest.manifest()
        inputs = [
            InputRef(
                dataset="chengjiao_xlsx",
                fetched_at=latest.fetched_at,
                content_hash=manifest.files[0].sha256 if manifest.files else None,
            )
        ]
    try:
        result = stage_xlsx(args.input, data_dir=data_dir, inputs=inputs or None)
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[xlsx stage] {exc}")
        return 1
    conserved = (
        result.sale_record_count == result.ordinary_residential_count + result.excluded_count
    )
    print(
        f"[xlsx stage] run_id={result.run_id} sale_record={result.sale_record_count} "
        f"普通住宅={result.ordinary_residential_count} 排除={result.excluded_count} "
        f"守恒={conserved}"
    )
    if inputs:
        print(
            f"[xlsx stage] 血缘: inputs=[{inputs[0].dataset}@{inputs[0].fetched_at} "
            f"sha256={str(inputs[0].content_hash)[:12]}…] "
            f"parser={result.sale_record_path and 'EXTFP1-C-1.0'}"
        )
    print(f"[xlsx stage] {SALE_RECORD_FILENAME} -> {result.sale_record_path}")
    print(f"[xlsx stage] {ORDINARY_FILENAME} -> {result.ordinary_residential_path}")
    print(f"[xlsx stage] 当前指针 -> {result.current_pointer}")
    return 0


def _cmd_xlsx_attributes_stage(args: argparse.Namespace) -> int:
    """excel-attribute-enrichment：从指定 staged run 只读派生属性标准化 v2 run。

    五列（楼层/装修/建成时间/电梯等）原文标准化为属性列（缺失语义按数据字典
    §1，不补造），写入不可变新 run 目录并切换 staged current 指针；源 run 与
    冻结版本目录零改动。质量摘要落 run 目录 attributes_quality.json。
    """
    from compsval.ingest.xlsx_stage import stage_attributes_run

    data_dir = args.data_dir or config.data_dir()
    try:
        result = stage_attributes_run(
            args.source_run_id,
            data_dir=data_dir,
            target_run_id=args.target_run_id,
        )
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[xlsx attributes-stage] {exc}")
        return 1
    coverage = {k: round(v, 4) for k, v in result.summary.coverage().items()}
    print(
        f"[xlsx attributes-stage] run_id={result.run_id} source={result.source_run_id}"
        f" sale_record={result.sale_record_count} 普通住宅={result.ordinary_residential_count}"
    )
    print(f"[xlsx attributes-stage] coverage={coverage}")
    print(f"[xlsx attributes-stage] quality -> {result.quality_json}")
    print(f"[xlsx attributes-stage] 当前指针 -> {result.current_pointer}")
    return 0


def _cmd_xlsx_quality(args: argparse.Namespace) -> int:
    """EXTFP1-E：从数据湖 staged 两表生成质量报告（MD + JSON）与回滚点。

    只读 `data_dir/staged/` 两表，统计守恒/字段质量/URL 分布/面积一致性，
    输出机器 JSON 与 Markdown（描述同一冻结数据）。git 基线由 `--git-baseline`
    提供或回退到当前 HEAD。
    """
    import subprocess

    from compsval.ingest.xlsx_quality import (
        build_xlsx_quality_report,
        load_staged_tables,
        write_xlsx_quality,
    )

    data_dir = args.data_dir or config.data_dir()
    try:
        sale_table, ordinary_table, manifests, run_id = load_staged_tables(data_dir)
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[xlsx quality] 无法读取 staged 表（先运行 compsval xlsx stage）：{exc}")
        return 1
    source_sha256 = None
    sale_manifest = manifests[0] if manifests else {}
    inputs_raw = sale_manifest.get("inputs")
    inputs_list = inputs_raw if isinstance(inputs_raw, list) else []
    if inputs_list and isinstance(inputs_list[0], dict):
        source_sha256 = str(inputs_list[0].get("content_hash") or "")
    if not source_sha256:
        notes = str(sale_manifest.get("notes") or "")
        if "sha256=" in notes:
            source_sha256 = notes.split("sha256=", 1)[1].strip() or None
    git_baseline = args.git_baseline
    if not git_baseline:
        try:
            completed = subprocess.run(
                ["git", "rev-parse", "HEAD"], capture_output=True, text=True, check=False
            )
            git_baseline = completed.stdout.strip() or None
        except FileNotFoundError:
            git_baseline = None
    try:
        report = build_xlsx_quality_report(
            sale_table,
            ordinary_table,
            source_sha256=source_sha256,
            git_baseline=git_baseline,
            staged_manifests=manifests,
        )
        md_path, json_path = write_xlsx_quality(report, data_dir=data_dir, out_json=args.out)
    except Exception as exc:  # noqa: BLE001 - user-facing failure
        print(f"[xlsx quality] {exc}")
        return 1
    c = report.counts
    print(
        f"[xlsx quality] run_id={run_id or 'UNKNOWN'} sale_record={c['sale_record_rows']} "
        f"普通住宅={c['ordinary_residential_rows']} 排除={c['excluded_rows']} "
        f"守恒={'通过' if c['conserved'] else '不通过'} "
        f"staged_tables={len(report.staged_tables)}"
    )
    print(f"[xlsx quality] MD -> {md_path}")
    print(f"[xlsx quality] JSON -> {json_path}")
    return 0


def _cmd_system_check(args: argparse.Namespace) -> int:
    """Run the offline quality gate: ruff, mypy, pytest.

    All checks run inside the current project environment. A non-zero exit from
    any check fails the gate.
    """
    root = Path(__file__).resolve().parents[2]
    steps: list[tuple[str, list[str]]] = [
        ("ruff", ["ruff", "check", "src", "tests"]),
        ("mypy", ["mypy", "src", "tests"]),
        ("pytest", ["pytest"]),
    ]
    failures: list[str] = []
    for name, cmd in steps:
        print(f"[system check] running {name} ...")
        try:
            completed = subprocess.run(cmd, cwd=root, capture_output=True, text=True)
        except FileNotFoundError:
            # Command not on PATH; for ruff/mypy prefer the uv-managed env.
            try:
                completed = subprocess.run(
                    ["uv", "run", *cmd], cwd=root, capture_output=True, text=True
                )
            except FileNotFoundError:
                print(f"[system check] {name}: executable not found (is uv on PATH?)")
                failures.append(name)
                continue
        if completed.stdout:
            sys.stdout.write(completed.stdout)
        if completed.stderr:
            sys.stderr.write(completed.stderr)
        if completed.returncode != 0:
            failures.append(name)
            print(f"[system check] {name}: FAILED (exit {completed.returncode})")
        else:
            print(f"[system check] {name}: ok")

    if failures:
        print(f"[system check] FAILED: {', '.join(failures)}")
        return 1
    print("[system check] all gates passed")
    return 0


def _snapshot_ref_by_id(data_dir: Path, snapshot_id: str) -> catalog.SnapshotRef | None:
    """Resolve a snapshot by its canonical id into a catalog SnapshotRef."""
    candidates: dict[str, catalog.SnapshotRef] = {}
    for ref in catalog.list_snapshots(data_dir):
        candidates[f"{ref.source}/{ref.dataset}@{ref.fetched_at}"] = ref
        candidates[f"{ref.source}-{ref.dataset}-{ref.fetched_at}"] = ref
    return candidates.get(snapshot_id)


def _cmd_data_stage(args: argparse.Namespace) -> int:
    """Re-derive staged + marts tables and the data-quality report from a snapshot.

    Command exit code reflects *completion* (did the stage run), not the
    business quality verdict (which lives in the report). Snapshot ids are the
    ``source/dataset@fetched_at`` forms shown by ``compsval catalog``.
    """
    data_dir = args.data_dir or config.data_dir()
    ref = _snapshot_ref_by_id(data_dir, args.snapshot)
    if ref is None:
        print(f"[data stage] no snapshot matching {args.snapshot!r} under {data_dir}")
        return 1
    result = data_stage(ref, data_dir=data_dir)
    print(
        f"[data stage] {args.snapshot}: sale_event={result.sale_event_path}"
        f" listing_event={result.listing_event_path}"
    )
    print(
        f"[data stage] marts: valid_sale={result.valid_sale_path}"
        f" valid_listing={result.valid_listing_path}"
    )
    print(
        f"[data stage] quality: markdown={result.quality_report_md}"
        f" json={result.quality_report_json}"
    )
    print(
        f"[data stage] summary: formal_pool={result.summary.formal_pool}"
        f" duplicates={result.summary.duplicate_flagged}"
        f" parking={result.summary.parking_flagged}"
    )
    return 0


def _cmd_data_marts_build(args: argparse.Namespace) -> int:
    """Build the combined multi-source marts (lianjia + fang_esf) for valuation.

    G3R-C: merges every participating raw snapshot into one ``valid_sale`` /
    ``valid_listing`` mart with cross-source dedup, then writes the merged
    data-quality report. Missing merge sources → non-zero exit with a clear
    message (no partial output).
    """
    data_dir = args.data_dir or config.data_dir()
    try:
        result = build_combined_marts(data_dir=data_dir)
    except Exception as exc:  # noqa: BLE001 - CLI surface for missing-input errors
        print(f"[data marts-build] {exc}")
        return 1
    print(f"[data marts-build] snapshots={len(result.snapshot_ids)}")
    print(f"[data marts-build] marts: valid_sale={result.valid_sale_path}")
    print(f"[data marts-build] marts: valid_listing={result.valid_listing_path}")
    print(f"[data marts-build] quality: markdown={result.quality_md}")
    print(f"[data marts-build] quality: json={result.quality_json}")
    print(
        f"[data marts-build] summary: formal_pool={result.summary.formal_pool}"
        f" duplicates={result.summary.duplicate_flagged}"
        f" parking={result.summary.parking_flagged}"
        f" cross_source_duplicates={result.cross_source_duplicates}"
        f" layout_backfilled={result.layout_backfilled}"
    )
    if result.ext_run_id is not None:
        print(
            f"[data marts-build] ext: run={result.ext_run_id}"
            f" input={result.ext_input_rows}"
            f" kept={result.ext_kept_rows}"
            f" unmatched_rows={result.ext_unmatched_rows}"
            f" unmatched_names={result.ext_unmatched_names}"
            f" attribute_matched={result.attribute_matched_rows}"
        )
    return 0


def _cmd_data_marts_enrich(args: argparse.Namespace) -> int:
    """excel-attribute-enrichment：身份键 join 把 staged v2 属性回填 valid_sale。

    显式输出到 ``--out``（不改写既有 marts/valid_sale.parquet；正式切换属
    基线确认后动作）。命中行带行级注记；无匹配留 None；命中率/冲突/前后
    覆盖率写入产物 manifest notes。
    """
    from compsval.ingest.attribute_enrich import enrich_attributes_mart

    data_dir = args.data_dir or config.data_dir()
    try:
        path, stats = enrich_attributes_mart(
            data_dir=data_dir,
            out_path=args.out,
            run_id=args.run_id,
        )
    except Exception as exc:  # noqa: BLE001 - CLI surface for missing-input errors
        print(f"[data marts-enrich-attributes] {exc}")
        return 1
    print(
        f"[data marts-enrich-attributes] run={stats.source_run_id}"
        f" matched={stats.matched}/{stats.rows_total}"
        f" conflict_keys={stats.conflict_keys}"
        f" unmapped_excel_rows={stats.excel_rows_unmapped_community}"
    )
    coverage_after = {k: round(v, 4) for k, v in stats.coverage_after.items()}
    print(f"[data marts-enrich-attributes] coverage_after={coverage_after}")
    print(f"[data marts-enrich-attributes] valid_sale(扩列) -> {path}")
    return 0


def _cmd_entities_build(args: argparse.Namespace) -> int:
    """Build the entity tables (community, alias, building, market_series, scope_policy).

     Derives ``data/entities/community.parquet``,
     ``data/entities/community_alias.parquet``, ``data/entities/building.parquet``,
     ``data/entities/market_series.parquet`` and the versioned scope policy list
     ``data/entities/scope_policy_v<version>.parquet`` (each row traceable to the
    名录/snapshot), plus their DerivedManifests. Prints the待人工确认 conflict
     checklist for the community_alias table. Does not touch raw snapshots or
     staged/marts tables.
    """
    data_dir = args.data_dir or config.data_dir()
    community_notes = (
        "WP5-A 骨架期：候选小区名录（房天下，SRC-005）转录 235 行；坐标/地址未知"
        "不虚构；链家全量权威数据待 WP5-B/E 回填"
    )
    path = entities_community.build_community_entity(data_dir=data_dir, notes=community_notes)
    community_table = pq.read_table(path)
    print(
        f"[entities build] community -> {path}"
        f" ({community_table.num_rows} rows; layer={entities_community.ENTITIES_LAYER})"
    )

    alias_notes = (
        "WP5-B：候选名录冲突清单 #1-8 名称类别名落表（community_alias，#9/#10 非名称类"
        "在待人工确认清单登记）；冲突不静默合并，进复核"
    )
    alias_path = entities_alias.build_alias_entity(data_dir=data_dir, notes=alias_notes)
    alias_table = pq.read_table(alias_path)
    print(
        f"[entities build] community_alias -> {alias_path}"
        f" ({alias_table.num_rows} rows; layer={entities_alias.ENTITIES_LAYER})"
    )
    print("[entities build] pending human-confirmation checklist (冲突清单 #1-10):")
    for conflict in entities_alias.pending_confirmation():
        print(
            f"  #{conflict.conflict_no} [{conflict.status.value}] "
            f"{conflict.title}\n      {conflict.action}"
        )

    building_notes = (
        "WP5-C：链家成交列表快照（SRC-007）楼栋弱实体匹配；低置信（LOW）进待复核"
        "清单不落表；未知字段用 UNKNOWN/None 不用 0"
    )
    try:
        building_path, low_rows = entities_building.build_building_entity(
            data_dir=data_dir, notes=building_notes
        )
    except FileNotFoundError as exc:
        print(f"[entities build] building skipped: {exc}")
    else:
        building_table = pq.read_table(building_path)
        print(
            f"[entities build] building -> {building_path}"
            f" ({building_table.num_rows} rows; layer={entities_community.ENTITIES_LAYER})"
        )
        if low_rows:
            print("[entities build] building 低置信待复核清单（LOW，不自动合并）:")
            for low in low_rows:
                print(
                    f"  {low.building_id} community={low.community_id}"
                    f" building_name={low.building_name} [{low.match_confidence.value}]\n"
                    f"      {low.source_ref}"
                )

    market_notes = (
        "WP5-D：58 同城板块均价（SRC-008，名录 §1.1）市场序列登记；板块聚合口径"
        "不冒充官方；聚合价不视为逐套成交；不做时间修正计算（归 WP6 VAL1-004）"
    )
    market_path = entities_market_series.build_market_series_entity(
        data_dir=data_dir, notes=market_notes
    )
    market_table = pq.read_table(market_path)
    print(
        f"[entities build] market_series -> {market_path}"
        f" ({market_table.num_rows} rows; layer={entities_community.ENTITIES_LAYER})"
    )

    scope_notes = (
        "WP5-F：ScopePolicy 适用范围判断（VAL1-001）从 community 权威表 + "
        "DATA-001 §4 冻结集合（11 个可实施 ID）派生范围清单；版本化输出不覆盖旧结果"
    )
    scope_path = valuation_scope.build_scope_policy(data_dir=data_dir, notes=scope_notes)
    scope_table = pq.read_table(scope_path)
    print(
        f"[entities build] scope_policy_v{valuation_scope.DEFAULT_RULE_VERSION} -> "
        f"{scope_path} ({scope_table.num_rows} rows; layer="
        f"{entities_community.ENTITIES_LAYER})"
    )
    decisions = scope_table.column("scope_decision").to_pylist()
    print(
        "[entities build] 范围清单分类："
        f"纳入={decisions.count('纳入')} 参考={decisions.count('参考')} "
        f"拒绝={decisions.count('拒绝')}"
    )
    return 0


def _cmd_estimate(args: argparse.Namespace) -> int:
    """WP7-B 端到端估值：subject JSON → 冻结估值 JSON（§10.3 包络）。

    stdout 只写机器可解析 JSON（成功为 success 包络，失败为 failure 包络，
    退出码 0/2/3/4/5 与 §10.4 一致）；日志与进度不写 stdout。非交互、无网络。

    formal 输出走受控启用路径（CX-WP9-02）：读取
    ``<data_dir>/release/release_decision.json``（RELEASE1-001 用户发布决定
    的运行载体）——记录缺失/无效/未发布一律保持候选/参考；CLI 不提供自由
    开关，启用只能经由该记录文件。
    """
    data_dir = args.data_dir or config.data_dir()
    if not args.subject.is_file():
        err = InvalidInputError(f"subject 文件不存在：{args.subject}")
        print(envelope_from_error("estimate", err).model_dump_json())
        return err.exit_code
    try:
        subject = SubjectProperty.model_validate_json(args.subject.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001 - 输入不合法（退出码 2）
        err = InvalidInputError(f"subject 解析失败：{exc}")
        print(envelope_from_error("estimate", err).model_dump_json())
        return err.exit_code
    as_of = None
    if args.as_of:
        try:
            as_of = date.fromisoformat(args.as_of)
        except ValueError:
            err = InvalidInputError(f"--as-of 格式非法：{args.as_of}（应为 YYYY-MM-DD）")
            print(envelope_from_error("estimate", err).model_dump_json())
            return err.exit_code
    release = valuation_estimate.load_release_decision(data_dir)
    try:
        outcome = valuation_estimate.run_estimate(
            subject=subject,
            as_of=as_of,
            data_dir=data_dir,
            out_root=args.out_dir,
            rule_version=args.rule_version,
            formal_release_enabled=release.enabled,
        )
    except CommandError as exc:
        print(envelope_from_error("estimate", exc).model_dump_json())
        return exc.exit_code
    except Exception as exc:  # noqa: BLE001 - 未预期内部错误（退出码 5）
        internal_err = InternalCommandError(f"未预期内部错误：{exc!r}")
        print(envelope_from_error("estimate", internal_err).model_dump_json())
        return internal_err.exit_code
    if release.enabled:
        outcome.envelope.add_warning(f"formal 输出已按发布决定记录启用：{release.detail}")
    elif release.recorded:
        outcome.envelope.add_warning(release.detail)
    print(outcome.envelope.model_dump_json(indent=2))
    return EXIT_OK


def _cmd_run_show(args: argparse.Namespace) -> int:
    """WP7-C 运行清单查询：compsval run show <run_id>（stdout 只写 §10.3 JSON）。"""
    data_dir = args.data_dir or config.data_dir()
    reports_root = args.out_dir or valuation_estimate.DEFAULT_REPORTS_ROOT
    try:
        envelope = reporting_run_show.show_run(
            run_id=args.run_id, data_dir=data_dir, reports_root=reports_root
        )
    except CommandError as exc:
        print(envelope_from_error("run show", exc).model_dump_json())
        return exc.exit_code
    except Exception as exc:  # noqa: BLE001 - 未预期内部错误（退出码 5）
        internal_err = InternalCommandError(f"未预期内部错误：{exc!r}")
        print(envelope_from_error("run show", internal_err).model_dump_json())
        return internal_err.exit_code
    print(envelope.model_dump_json(indent=2))
    return EXIT_OK


def _cmd_report_build(args: argparse.Namespace) -> int:
    """WP7-C Markdown 报告生成：compsval report build --valuation <id>（§11.2 十二节）。

    从冻结估值 JSON + 中间结果表只读生成报告；不修改冻结 JSON。stdout 只写
    §10.3 JSON。
    """
    data_dir = args.data_dir or config.data_dir()
    reports_root = args.out_dir or valuation_estimate.DEFAULT_REPORTS_ROOT
    try:
        result = reporting_markdown.build_report_markdown(
            run_id=args.valuation, data_dir=data_dir, reports_root=reports_root
        )
    except CommandError as exc:
        print(envelope_from_error("report build", exc).model_dump_json())
        return exc.exit_code
    except Exception as exc:  # noqa: BLE001 - 未预期内部错误（退出码 5）
        internal_err = InternalCommandError(f"未预期内部错误：{exc!r}")
        print(envelope_from_error("report build", internal_err).model_dump_json())
        return internal_err.exit_code
    report_path = reporting_markdown.report_path_for(args.valuation, reports_root)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    work = report_path.with_name(report_path.name + ".incomplete")
    work.write_text(result.markdown, encoding="utf-8")
    work.replace(report_path)
    # WP8-B 版本治理回填（RV-WP7-C-02 F3）：包络 data_version/rule_version 从 run 表填充
    data_version, rule_version = reporting_run_show.run_versions(
        run_id=args.valuation, data_dir=data_dir
    )
    envelope = OutputEnvelope(
        command="report build",
        business_status=result.business_status,
        run_id=args.valuation,
        data_version=data_version,
        rule_version=rule_version,
        artifacts=[str(report_path)],
    )
    print(envelope.model_dump_json(indent=2))
    return EXIT_OK


def _cmd_review_apply(args: argparse.Namespace) -> int:
    """WP7-D 复核留痕：compsval review apply --valuation <id> --input <json>。

    校验估值结果存在（退出码 3）→ 只追加 review_event（WP6-F）→ 输出新版本
    引用（review_id + 声明不覆盖自动结果）。stdout 只写 §10.3 JSON。
    """
    data_dir = args.data_dir or config.data_dir()
    if not args.input.is_file():
        err = InvalidInputError(f"复核输入文件不存在：{args.input}")
        print(envelope_from_error("review apply", err).model_dump_json())
        return err.exit_code
    try:
        payload = json.loads(args.input.read_text(encoding="utf-8-sig"))
    except (json.JSONDecodeError, OSError) as exc:
        err = InvalidInputError(f"复核 JSON 解析失败：{exc}")
        print(envelope_from_error("review apply", err).model_dump_json())
        return err.exit_code
    if not isinstance(payload, dict):
        err = InvalidInputError("复核 JSON 必须为对象（技术方案 §11.1）")
        print(envelope_from_error("review apply", err).model_dump_json())
        return err.exit_code
    try:
        review_path, events = valuation_review_apply.apply_review_for_run(
            run_id=args.valuation, data_dir=data_dir, input_payload=payload
        )
    except CommandError as exc:
        print(envelope_from_error("review apply", exc).model_dump_json())
        return exc.exit_code
    except Exception as exc:  # noqa: BLE001 - 未预期内部错误（退出码 5）
        internal_err = InternalCommandError(f"未预期内部错误：{exc!r}")
        print(envelope_from_error("review apply", internal_err).model_dump_json())
        return internal_err.exit_code
    # WP8-B 版本治理回填（RV-WP7-D-01 F4）：包络 data_version/rule_version 从 run 表填充
    data_version, rule_version = reporting_run_show.run_versions(
        run_id=args.valuation, data_dir=data_dir
    )
    envelope = OutputEnvelope(
        command="review apply",
        run_id=args.valuation,
        data_version=data_version,
        rule_version=rule_version,
        result={
            "result_id": events[0].result_id,
            "review_ids": [e.review_id for e in events],
            "statement": "复核只追加留痕，不覆盖自动估值结果（自动结果原样保留）",
        },
        artifacts=[str(review_path)],
    )
    print(envelope.model_dump_json(indent=2))
    return EXIT_OK


def _cmd_backtest_run(args: argparse.Namespace) -> int:
    """WP8-A 滚动历史回放：compsval backtest run --config <yaml>（§13.3/§10.3/§10.4）。

    stdout 只写机器可解析 JSON（§10.3 包络）；缺配置/坏配置 → 退出码 2，
    必要数据表缺失 → 3，未预期内部错误 → 5。非交互、无网络。
    """
    data_dir = args.data_dir or config.data_dir()
    try:
        bt_config = valuation_backtest.load_backtest_config(args.config)
    except CommandError as exc:
        print(envelope_from_error("backtest run", exc).model_dump_json())
        return exc.exit_code
    try:
        outcome = valuation_backtest.run_backtest(bt_config, data_dir=data_dir)
    except CommandError as exc:
        print(envelope_from_error("backtest run", exc).model_dump_json())
        return exc.exit_code
    except Exception as exc:  # noqa: BLE001 - 未预期内部错误（退出码 5）
        internal_err = InternalCommandError(f"未预期内部错误：{exc!r}")
        print(envelope_from_error("backtest run", internal_err).model_dump_json())
        return internal_err.exit_code

    metrics = {
        metric: value
        for metric, value in zip(
            outcome.metrics.column("metric").to_pylist(),
            outcome.metrics.column("value").to_pylist(),
            strict=True,
        )
    }
    n_estimated = int(metrics.get("n_estimated") or 0)
    n_skipped = int(metrics.get("n_skipped") or 0)
    business_status = "参考" if n_estimated > 0 else "信息不足"
    warnings: list[str] = []
    if n_estimated == 0:
        warnings.append("无任何回放校验样本，回放覆盖为 0，不得据此校准数值门槛")
    if n_skipped > 0:
        warnings.append(f"{n_skipped} 个目标成交因未匹配小区/缺必要字段被跳过（如实报告）")
    envelope = OutputEnvelope(
        command="backtest run",
        business_status=business_status,
        run_id=outcome.run_id,
        data_version=outcome.data_version,
        rule_version=bt_config.rule_version,
        result={
            "run_id": outcome.run_id,
            "metrics": metrics,
            "detail_path": str(outcome.detail_path),
            "metrics_path": str(outcome.metrics_path),
            "run_manifest_path": str(outcome.run_manifest_path),
        },
        warnings=warnings,
        artifacts=[
            str(outcome.detail_path),
            str(outcome.metrics_path),
            str(outcome.run_manifest_path),
        ],
    )
    print(envelope.model_dump_json(indent=2))
    return EXIT_OK


def _cmd_backtest_report(args: argparse.Namespace) -> int:
    """WP8-B 可复现回放报告：compsval backtest report --run <run_id>（§13.3/§12.2）。

    从回放产物只读生成 Markdown + JSON 报告（04-校验/backtest_reports/）；
    run_id 与回放运行清单不一致 → 退出码 2，回放产物缺失 → 3。stdout 只写
    §10.3 包络。
    """
    data_dir = args.data_dir or config.data_dir()
    reports_root = args.out_dir or reporting_backtest_report.DEFAULT_BACKTEST_REPORTS_ROOT
    try:
        result = reporting_backtest_report.build_backtest_report(
            run_id=args.run,
            data_dir=data_dir,
            out_root=reports_root,
            backtest_dir=getattr(args, "backtest_dir", None),
        )
    except CommandError as exc:
        print(envelope_from_error("backtest report", exc).model_dump_json())
        return exc.exit_code
    except Exception as exc:  # noqa: BLE001 - 未预期内部错误（退出码 5）
        internal_err = InternalCommandError(f"未预期内部错误：{exc!r}")
        print(envelope_from_error("backtest report", internal_err).model_dump_json())
        return internal_err.exit_code
    envelope = OutputEnvelope(
        command="backtest report",
        run_id=result.run_id,
        data_version=result.data_version,
        rule_version=result.rule_version,
        result={
            "run_id": result.run_id,
            "report_markdown_path": str(result.markdown_path),
            "report_json_path": str(result.json_path),
        },
        artifacts=[str(result.markdown_path), str(result.json_path)],
    )
    print(envelope.model_dump_json(indent=2))
    return EXIT_OK


def _cmd_shadow_register(args: argparse.Namespace) -> int:
    """WP9-A 影子标的登记：compsval shadow register --subject <json>（SHADOW-001）。

    复用 WP7-B ``run_estimate`` 冻结估值并把 frozen 结果登记到影子追踪表
    （只追加不改写：同 run_id 已登记 → 幂等返回既有行）。stdout 只写 §10.3
    包络；退出码 0/2/3/4/5。
    """
    data_dir = args.data_dir or config.data_dir()
    if not args.subject.is_file():
        err = InvalidInputError(f"subject 文件不存在：{args.subject}")
        print(envelope_from_error("shadow register", err).model_dump_json())
        return err.exit_code
    try:
        subject = SubjectProperty.model_validate_json(args.subject.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001 - 输入不合法（退出码 2）
        err = InvalidInputError(f"subject 解析失败：{exc}")
        print(envelope_from_error("shadow register", err).model_dump_json())
        return err.exit_code
    try:
        outcome = valuation_shadow.register_subject(
            subject=subject,
            data_dir=data_dir,
            out_root=args.out_dir,
            rule_version=args.rule_version,
            notes=args.note,
        )
    except CommandError as exc:
        print(envelope_from_error("shadow register", exc).model_dump_json())
        return exc.exit_code
    except Exception as exc:  # noqa: BLE001 - 未预期内部错误（退出码 5）
        internal_err = InternalCommandError(f"未预期内部错误：{exc!r}")
        print(envelope_from_error("shadow register", internal_err).model_dump_json())
        return internal_err.exit_code
    envelope = OutputEnvelope(
        command="shadow register",
        business_status=outcome.frozen_business_status,
        run_id=outcome.run_id,
        data_version=outcome.data_version,
        rule_version=outcome.rule_version,
        result={
            "subject_id": outcome.subject_id,
            "run_id": outcome.run_id,
            "estimate_path": str(outcome.estimate_path),
            "frozen_business_status": outcome.frozen_business_status,
            "duplicated": outcome.duplicated,
        },
        artifacts=[str(outcome.estimate_path), str(outcome.track_path)],
    )
    if outcome.duplicated:
        envelope.add_warning(f"run {outcome.run_id} 已登记，返回既有追踪行（追踪表只读不改写）")
    print(envelope.model_dump_json(indent=2))
    return EXIT_OK


def _cmd_shadow_backfill(args: argparse.Namespace) -> int:
    """WP9-A 后续成交回填：compsval shadow backfill（时间外，不挑选样本）。

    对每个影子标的，用 ``估值时点 < sale_date <= tracking_cutoff`` 的同小区
    成交计算 APE/区间命中，全量重建 ``shadow_followup``（可复现）。stdout
    只写 §10.3 包络；退出码 0/2/3/4/5。
    """
    data_dir = args.data_dir or config.data_dir()
    cutoff: date | None = None
    if args.tracking_cutoff:
        try:
            cutoff = date.fromisoformat(args.tracking_cutoff)
        except ValueError:
            err = InvalidInputError(
                f"--tracking-cutoff 格式非法：{args.tracking_cutoff}（应为 YYYY-MM-DD）"
            )
            print(envelope_from_error("shadow backfill", err).model_dump_json())
            return err.exit_code
    try:
        outcome = valuation_shadow.backfill_followups(data_dir=data_dir, tracking_cutoff=cutoff)
    except CommandError as exc:
        print(envelope_from_error("shadow backfill", exc).model_dump_json())
        return exc.exit_code
    except Exception as exc:  # noqa: BLE001 - 未预期内部错误（退出码 5）
        internal_err = InternalCommandError(f"未预期内部错误：{exc!r}")
        print(envelope_from_error("shadow backfill", internal_err).model_dump_json())
        return internal_err.exit_code
    envelope = OutputEnvelope(
        command="shadow backfill",
        run_id=None,
        data_version=outcome.data_version,
        result={
            "n_subjects": outcome.n_subjects,
            "n_followup_sales": outcome.n_followup,
            "tracking_cutoff": outcome.tracking_cutoff.isoformat(),
            "followup_path": str(outcome.followup_path),
        },
        artifacts=[str(outcome.followup_path)],
    )
    if outcome.n_followup == 0:
        envelope.add_warning("无后续成交可追踪（如实报告，不以虚构结果宣称 G4 通过）")
    print(envelope.model_dump_json(indent=2))
    return EXIT_OK


def _cmd_shadow_monitor(args: argparse.Namespace) -> int:
    """WP9-A 影子误差监控：compsval shadow monitor（近期滚动窗口 + 数据新鲜度）。

    按 README §7.2 触发条件（误差扩大/区间失准/数据中断）输出预警；样本不足
    如实标注不判定。stdout 只写 §10.3 包络；退出码 0/2/3/4/5。
    """
    data_dir = args.data_dir or config.data_dir()
    as_of: date | None = None
    if args.as_of:
        try:
            as_of = date.fromisoformat(args.as_of)
        except ValueError:
            err = InvalidInputError(f"--as-of 格式非法：{args.as_of}（应为 YYYY-MM-DD）")
            print(envelope_from_error("shadow monitor", err).model_dump_json())
            return err.exit_code
    monitor_config = valuation_shadow.ShadowMonitorConfig(
        window_days=args.window_days,
        baseline_ape_median=args.baseline_ape_median,
        baseline_range_coverage=args.baseline_range_coverage,
        stale_days=args.stale_days,
        as_of=as_of,
    )
    try:
        report = valuation_shadow.monitor(data_dir=data_dir, config=monitor_config)
    except CommandError as exc:
        print(envelope_from_error("shadow monitor", exc).model_dump_json())
        return exc.exit_code
    except Exception as exc:  # noqa: BLE001 - 未预期内部错误（退出码 5）
        internal_err = InternalCommandError(f"未预期内部错误：{exc!r}")
        print(envelope_from_error("shadow monitor", internal_err).model_dump_json())
        return internal_err.exit_code
    envelope = OutputEnvelope(
        command="shadow monitor",
        run_id=None,
        result={
            "as_of": report.as_of.isoformat(),
            "window_days": monitor_config.window_days,
            "window_metrics": report.window_metrics,
            "signed_metrics": report.signed_metrics,
            "freshness": report.freshness,
            "triggers": report.triggers,
            "subjects": report.subjects,
        },
        warnings=[t["label"] for t in report.triggers],
    )
    print(envelope.model_dump_json(indent=2))
    return EXIT_OK


def _cmd_valuation_build(args: argparse.Namespace) -> int:
    """Build WP6-A 估值中间结果表（subject_property/run/comp_candidate）。

    Reads a subject JSON (数据字典 §3.9), fixes valuation_date/data_cutoff,
    retrieves the candidate pool from ``valid_sale`` (数据截点之前的成交；排除
    非住宅/车位、明显异常、缺必要字段、community_id 未匹配，reason 全量留痕),
    and writes the three intermediate tables under ``data/valuation/`` with
    DerivedManifests. Prints入选/排除 statistics for inspection.
    """
    data_dir = args.data_dir or config.data_dir()
    if not args.subject.is_file():
        print(f"[valuation build] subject file not found: {args.subject}")
        return 1
    try:
        subject = SubjectProperty.model_validate_json(args.subject.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001 - user-facing validation error
        print(f"[valuation build] invalid subject: {exc}")
        return 1
    try:
        result = valuation_candidate.build_valuation(subject, data_dir=data_dir)
    except FileNotFoundError as exc:
        print(f"[valuation build] {exc}")
        return 1
    selected = sum(1 for c in result.candidates if c.selected)
    excluded = len(result.candidates) - selected
    print(f"[valuation build] run {result.run.run_id}")
    print(f"  subject_property -> {result.subject_path}")
    print(f"  valuation_run -> {result.run_path}")
    print(
        f"  comp_candidate -> {result.candidate_path}"
        f" ({len(result.candidates)} rows; selected={selected} excluded={excluded})"
    )
    for reason, count in Counter(
        c.reason for c in result.candidates if not c.selected
    ).most_common():
        print(f"  - {reason}: {count}")
    return 0


def _cmd_valuation_tier(args: argparse.Namespace) -> int:
    """Build WP6-B 可比层级：给 comp_candidate 填 tier/similarity 并写竞争关系清单。

    Reads the WP6-A ``data/valuation/comp_candidate.parquet``, classifies each
    selected candidate into可比层级 A-E（一次只放宽一个主要条件）、计算相似度，
    and discovers 候选竞争小区关系（同板块，待人工确认）清单. 原子写回
    comp_candidate.parquet + 写 competitive_relations.parquet + DerivedManifests.
    """
    data_dir = args.data_dir or config.data_dir()
    if not args.subject.is_file():
        print(f"[valuation tier] subject file not found: {args.subject}")
        return 1
    try:
        subject = SubjectProperty.model_validate_json(args.subject.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001 - user-facing validation error
        print(f"[valuation tier] invalid subject: {exc}")
        return 1
    valid_sale_path = data_dir / MARTS_LAYER / VALID_SALE_FILENAME
    community_path = (
        data_dir / entities_community.ENTITIES_LAYER / entities_community.COMMUNITY_FILENAME
    )
    try:
        valid_sale = pq.read_table(valid_sale_path)
        communities = pq.read_table(community_path)
    except FileNotFoundError as exc:
        print(f"[valuation tier] {exc}")
        return 1
    inputs = []
    manifest_path = valid_sale_path.with_suffix(".manifest.json")
    if manifest_path.is_file():
        inputs = list(read_derived_manifest(valid_sale_path).inputs)
    try:
        result = valuation_comparable.apply_tiers_to_candidate_table(
            data_dir=data_dir,
            subject=subject,
            valid_sale=valid_sale,
            communities=communities,
            input_refs=inputs,
        )
    except FileNotFoundError as exc:
        print(f"[valuation tier] {exc}")
        return 1
    tiered = [c for c in result.candidates if c.tier is not None]
    print(
        f"[valuation tier] comp_candidate -> {result.candidate_path}"
        f" ({len(result.candidates)} rows; tiered={len(tiered)})"
    )
    tiers_used = sorted({t for c in tiered if (t := c.tier) is not None})
    for tier in tiers_used:
        count = sum(1 for c in tiered if c.tier == tier)
        first = next(c for c in tiered if c.tier == tier)
        print(f"  - {first.reason}: {count}")
    pending = [r for r in result.relations if not r.confirmed]
    print(
        f"[valuation tier] competitive_relations -> {result.relations_path}"
        f" ({len(result.relations)} rows; pending human-confirm={len(pending)})"
    )
    return 0


def _cmd_valuation_time(args: argparse.Namespace) -> int:
    """Build WP6-C 时间修正：为可比候选写 comp_adjustment（adjustment_type=时间）。

    Reads the WP6-B ``data/valuation/comp_candidate.parquet`` selected
    comparables, resolves a time adjustment for each (only data available at the
    valuation date; degrade when no reliable series), and writes
    ``data/valuation/comp_adjustment.parquet`` + DerivedManifest.
    """
    data_dir = args.data_dir or config.data_dir()
    if not args.subject.is_file():
        print(f"[valuation time] subject file not found: {args.subject}")
        return 1
    try:
        subject = SubjectProperty.model_validate_json(args.subject.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001 - user-facing validation error
        print(f"[valuation time] invalid subject: {exc}")
        return 1
    valid_sale_path = data_dir / MARTS_LAYER / VALID_SALE_FILENAME
    community_path = (
        data_dir / entities_community.ENTITIES_LAYER / entities_community.COMMUNITY_FILENAME
    )
    market_series_path = data_dir / entities_community.ENTITIES_LAYER / "market_series.parquet"
    try:
        valid_sale = pq.read_table(valid_sale_path)
        communities = pq.read_table(community_path)
        market_series = pq.read_table(market_series_path)
    except FileNotFoundError as exc:
        print(f"[valuation time] {exc}")
        return 1
    inputs = []
    manifest_path = valid_sale_path.with_suffix(".manifest.json")
    if manifest_path.is_file():
        inputs = list(read_derived_manifest(valid_sale_path).inputs)
    try:
        result = valuation_time_adjustment.apply_time_adjustments(
            data_dir=data_dir,
            subject=subject,
            valid_sale=valid_sale,
            market_series=market_series,
            communities=communities,
            input_refs=inputs,
        )
    except FileNotFoundError as exc:
        print(f"[valuation time] {exc}")
        return 1
    print(
        f"[valuation time] comp_adjustment -> {result.adjustment_path}"
        f" ({len(result.adjustments)} rows; adjustment_type=时间)"
    )
    by_strength: Counter[str] = Counter(a.evidence_strength.value for a in result.adjustments)
    for values in sorted(by_strength):
        print(f"  - 证据强度 {values}: {by_strength[values]}")
    degraded = [a for a in result.adjustments if a.amount is None]
    if degraded:
        print(f"  - 降级（无可靠序列，amount=None）: {len(degraded)}")
    return 0


def _cmd_valuation_diff(args: argparse.Namespace) -> int:
    """Build WP6-D 房源差异：为可比候选写 comp_adjustment（adjustment_type=差异）。

    Reads the WP6-B ``data/valuation/comp_candidate.parquet`` selected
    comparables, resolves a six-dimension property difference against the
    subject (no fabricated ratios without market evidence), and appends
    ``adjustment_type=差异`` rows to the existing ``comp_adjustment.parquet``.
    """
    data_dir = args.data_dir or config.data_dir()
    if not args.subject.is_file():
        print(f"[valuation diff] subject file not found: {args.subject}")
        return 1
    try:
        subject = SubjectProperty.model_validate_json(args.subject.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001 - user-facing validation error
        print(f"[valuation diff] invalid subject: {exc}")
        return 1
    valid_sale_path = data_dir / MARTS_LAYER / VALID_SALE_FILENAME
    buildings_path = (
        data_dir / entities_community.ENTITIES_LAYER / entities_building.BUILDING_FILENAME
    )
    try:
        valid_sale = pq.read_table(valid_sale_path)
        buildings = pq.read_table(buildings_path)
    except FileNotFoundError as exc:
        print(f"[valuation diff] {exc}")
        return 1
    inputs = []
    manifest_path = valid_sale_path.with_suffix(".manifest.json")
    if manifest_path.is_file():
        inputs = list(read_derived_manifest(valid_sale_path).inputs)
    try:
        result = valuation_difference.apply_difference_adjustments(
            data_dir=data_dir,
            subject=subject,
            valid_sale=valid_sale,
            buildings=buildings,
            input_refs=inputs,
        )
    except FileNotFoundError as exc:
        print(f"[valuation diff] {exc}")
        return 1
    print(
        f"[valuation diff] comp_adjustment -> {result.adjustment_path}"
        f" (+{len(result.adjustments)} rows; adjustment_type=差异)"
    )
    by_feature: Counter[str] = Counter(a.feature for a in result.adjustments)
    for feature in sorted(by_feature):
        print(f"  - {feature}: {by_feature[feature]}")
    degraded = [a for a in result.adjustments if a.factor is None]
    if degraded:
        print(f"  - 降级（无市场证据，factor=None）: {len(degraded)}")
    return 0


def _cmd_valuation_aggregate(args: argparse.Namespace) -> int:
    """Build WP6-E 中心/区间/可信度/输出状态：写 valuation_result。

    Reads the WP6-B ``comp_candidate.parquet`` selected comparables and the
    WP6-C/D ``comp_adjustment.parquet`` adjustments, aggregates via similarity
    weighted median + weighted quantiles, classes confidence (high/medium/low/
    insufficient) with per-factor evidence, and decides output status
    (候选/参考/正式; formal only when the release gate is enabled). Writes
    ``valuation_result.parquet`` + DerivedManifest.
    """
    data_dir = args.data_dir or config.data_dir()
    if not args.subject.is_file():
        print(f"[valuation aggregate] subject file not found: {args.subject}")
        return 1
    try:
        subject = SubjectProperty.model_validate_json(args.subject.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001 - user-facing validation error
        print(f"[valuation aggregate] invalid subject: {exc}")
        return 1
    valid_sale_path = data_dir / MARTS_LAYER / VALID_SALE_FILENAME
    candidate_path = (
        data_dir / valuation_candidate.VALUATION_LAYER / valuation_candidate.COMP_CANDIDATE_FILENAME
    )
    try:
        valid_sale = pq.read_table(valid_sale_path)
    except FileNotFoundError as exc:
        print(f"[valuation aggregate] {exc}")
        return 1
    inputs = []
    if candidate_path.is_file():
        manifest_path = candidate_path.with_suffix(".manifest.json")
        if manifest_path.is_file():
            inputs = list(read_derived_manifest(candidate_path).inputs)
    try:
        result = valuation_aggregation.apply_aggregation(
            data_dir=data_dir,
            subject=subject,
            valid_sale=valid_sale,
            input_refs=inputs,
        )
    except ValueError as exc:
        print(f"[valuation aggregate] {exc}")
        return 1
    print(f"[valuation aggregate] valuation_result -> {result.result_path}")
    r = result.result
    print(f"  run={r.run_id} subject={r.subject_id}")
    print(f"  center={r.center} 范围=[{r.range_lower}, {r.range_upper}]")
    print(f"  可信度={r.confidence.value} 状态={r.status.value}")
    print(
        f"  可比={result.n_comps} 有效样本={result.effective_samples:.1f} 规则版本={r.rule_version}"
    )
    print(f"  reason: {r.reason}")
    return 0


def _cmd_valuation_review(args: argparse.Namespace) -> int:
    """Append 一条人工复核留痕：写 review_event（只追加，不覆盖自动结果，WP6-F）。"""
    data_dir = args.data_dir or config.data_dir()
    if not args.input.is_file():
        print(f"[valuation review] input file not found: {args.input}")
        return 1
    try:
        event = valuation_review.ReviewEventInput(
            **json.loads(args.input.read_text(encoding="utf-8"))
        )
    except Exception as exc:
        print(f"[valuation review] invalid review input: {exc}")
        return 1
    try:
        review_path, assigned = valuation_review.append_review_events(
            data_dir=data_dir,
            events=[event.to_review_event()],
            notes=f"WP6-F: 复核留痕 {event.action}",
        )
    except valuation_review.ReviewError as exc:
        print(f"[valuation review] {exc}")
        return 1
    print(f"[valuation review] review_event -> {review_path}")
    for ev in assigned:
        print(f"  review_id={ev.review_id}  action={ev.action.value}  judgment={ev.judgment.value}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="compsval",
        description="ExampleCity target-district residential valuation engine (skeleton).",
    )
    subparsers = parser.add_subparsers(dest="command")

    subparsers.add_parser("version", help="print the package version").set_defaults(
        func=_cmd_version
    )

    catalog_cmd = subparsers.add_parser(
        "catalog", help="list the complete raw snapshots for each dataset"
    )
    catalog_cmd.add_argument("--data-dir", type=Path)
    catalog_cmd.set_defaults(func=_cmd_catalog)

    estimate_cmd = subparsers.add_parser(
        "estimate",
        help=(
            "execute an end-to-end valuation: subject JSON -> frozen estimate "
            "JSON + unified envelope (REP-001/WP7-B; stdout writes JSON only)"
        ),
    )
    estimate_cmd.add_argument(
        "--subject",
        required=True,
        type=Path,
        help="subject JSON file (数据字典 §3.9)",
    )
    estimate_cmd.add_argument(
        "--as-of",
        help="valuation date YYYY-MM-DD (must equal subject.valuation_date)",
    )
    estimate_cmd.add_argument(
        "--rule-version",
        default="1.0",
        help=(
            "valuation rule version (default 1.0; 1.1 = interval-calibrated "
            "chain, requires the calibration config in <data-dir>/rules/)"
        ),
    )
    estimate_cmd.add_argument("--data-dir", type=Path)
    estimate_cmd.add_argument(
        "--out-dir",
        type=Path,
        help="frozen estimate output root (default: project 05-估值报告/)",
    )
    estimate_cmd.set_defaults(func=_cmd_estimate)

    run_cmd = subparsers.add_parser("run", help="inspect a valuation run manifest (WP7-C)")
    run_subparsers = run_cmd.add_subparsers(dest="run_command")
    run_show_cmd = run_subparsers.add_parser(
        "show",
        help=(
            "show the run manifest: versions, parameters, run time and product "
            "paths (WP7-C; stdout writes JSON only)"
        ),
    )
    run_show_cmd.add_argument(
        "run_id", help="run id (e.g. RUN-SUBJ-REAL-001-20260721-20260721-v1.0)"
    )
    run_show_cmd.add_argument("--data-dir", type=Path)
    run_show_cmd.add_argument(
        "--out-dir",
        type=Path,
        help="frozen estimate/report root (default: project 05-估值报告/)",
    )
    run_show_cmd.set_defaults(func=_cmd_run_show)
    run_show_cmd.set_defaults(run_command="show")
    run_cmd.set_defaults(func=_cmd_run_show)

    report_cmd = subparsers.add_parser("report", help="build valuation reports (WP7-C)")
    report_subparsers = report_cmd.add_subparsers(dest="report_command")
    report_build_cmd = report_subparsers.add_parser(
        "build",
        help=(
            "build the 12-section Markdown report from the frozen estimate "
            "(§11.2, WP7-C; read-only, never rewrites the frozen JSON)"
        ),
    )
    report_build_cmd.add_argument("--valuation", required=True, help="valuation run id")
    report_build_cmd.add_argument("--data-dir", type=Path)
    report_build_cmd.add_argument(
        "--out-dir",
        type=Path,
        help="frozen estimate/report root (default: project 05-估值报告/)",
    )
    report_build_cmd.set_defaults(func=_cmd_report_build)
    report_build_cmd.set_defaults(report_command="build")
    report_cmd.set_defaults(func=_cmd_report_build)

    review_cmd = subparsers.add_parser(
        "review",
        help="append a human review event (WP7-D; official §10.2 contract)",
    )
    review_subparsers = review_cmd.add_subparsers(dest="review_command")
    review_apply_cmd = review_subparsers.add_parser(
        "apply",
        help=(
            "append a review event for a valuation run: validates the frozen "
            "result, appends review_event (append-only, never overwrites), "
            "returns new event refs (WP7-D, §11.1)"
        ),
    )
    review_apply_cmd.add_argument(
        "--valuation", required=True, help="valuation run id (must exist in valuation_result)"
    )
    review_apply_cmd.add_argument(
        "--input",
        required=True,
        type=Path,
        help="review JSON contract file (技术方案 §11.1; result_id resolved from --valuation)",
    )
    review_apply_cmd.add_argument("--data-dir", type=Path)
    review_apply_cmd.set_defaults(func=_cmd_review_apply)
    review_apply_cmd.set_defaults(review_command="apply")
    review_cmd.set_defaults(func=_cmd_review_apply)

    backtest_cmd = subparsers.add_parser(
        "backtest", help="rolling out-of-time historical replay (BT-001/WP8-A)"
    )
    backtest_subparsers = backtest_cmd.add_subparsers(dest="backtest_command")
    backtest_run_cmd = backtest_subparsers.add_parser(
        "run",
        help=(
            "run a rolling out-of-time replay: time-respecting splits, simple "
            "same-community median baseline, §13.3 metrics (§13.3, WP8-A; "
            "stdout writes JSON only)"
        ),
    )
    backtest_run_cmd.add_argument(
        "--config",
        required=True,
        type=Path,
        help=(
            "backtest YAML config file (rule_version / replay_dates / "
            "baseline_window_months / high_quantile)"
        ),
    )
    backtest_run_cmd.add_argument("--data-dir", type=Path)
    backtest_run_cmd.set_defaults(func=_cmd_backtest_run)
    backtest_run_cmd.set_defaults(backtest_command="run")
    backtest_cmd.set_defaults(func=_cmd_backtest_run)

    backtest_report_cmd = backtest_subparsers.add_parser(
        "report",
        help=(
            "build the reproducible replay report (Markdown + JSON, §13.3/"
            "§12.2, WP8-B; read-only from replay products, never rewrites them)"
        ),
    )
    backtest_report_cmd.add_argument(
        "--run", required=True, help="backtest run id (must match run_manifest.json)"
    )
    backtest_report_cmd.add_argument("--data-dir", type=Path)
    backtest_report_cmd.add_argument(
        "--backtest-dir",
        type=Path,
        help=(
            "replay products dir override (default: <data-dir>/backtest; e.g. "
            "data/backtest-exp/<policy> for aggregator experiment runs)"
        ),
    )
    backtest_report_cmd.add_argument(
        "--out-dir",
        type=Path,
        help="replay report root (default: project 04-校验/backtest_reports/)",
    )
    backtest_report_cmd.set_defaults(func=_cmd_backtest_report)
    backtest_report_cmd.set_defaults(backtest_command="report")
    backtest_cmd.set_defaults(func=_cmd_backtest_report)

    shadow_cmd = subparsers.add_parser(
        "shadow", help="shadow-run infrastructure (SHADOW-001/WP9-A)"
    )
    shadow_subparsers = shadow_cmd.add_subparsers(dest="shadow_command")
    shadow_register_cmd = shadow_subparsers.add_parser(
        "register",
        help=(
            "register one shadow subject: freeze an estimate via the WP7 "
            "estimate chain and record the frozen result in shadow_track "
            "(append-only, never overwrites; SHADOW-001/WP9-A)"
        ),
    )
    shadow_register_cmd.add_argument(
        "--subject", required=True, type=Path, help="subject JSON file (数据字典 §3.9)"
    )
    shadow_register_cmd.add_argument("--data-dir", type=Path)
    shadow_register_cmd.add_argument(
        "--out-dir",
        type=Path,
        help="frozen estimate output root (default: project 05-估值报告/)",
    )
    shadow_register_cmd.add_argument(
        "--rule-version", default="1.0", help="valuation rule version (default 1.0)"
    )
    shadow_register_cmd.add_argument(
        "--note", help="optional provenance note for this shadow subject"
    )
    shadow_register_cmd.set_defaults(func=_cmd_shadow_register)
    shadow_register_cmd.set_defaults(shadow_command="register")
    shadow_cmd.set_defaults(func=_cmd_shadow_register)

    shadow_backfill_cmd = shadow_subparsers.add_parser(
        "backfill",
        help=(
            "backfill follow-up actual sales for all registered shadow subjects "
            "(out-of-time: valuation_date < sale_date <= tracking_cutoff, same "
            "community; rebuilds shadow_followup reproducibly; SHADOW-001/WP9-A)"
        ),
    )
    shadow_backfill_cmd.add_argument(
        "--tracking-cutoff",
        help="tracking data cutoff YYYY-MM-DD (default: latest valid_sale sale_date)",
    )
    shadow_backfill_cmd.add_argument("--data-dir", type=Path)
    shadow_backfill_cmd.set_defaults(func=_cmd_shadow_backfill)
    shadow_backfill_cmd.set_defaults(shadow_command="backfill")
    shadow_cmd.set_defaults(func=_cmd_shadow_backfill)

    shadow_monitor_cmd = shadow_subparsers.add_parser(
        "monitor",
        help=(
            "report rolling-window error metrics and data freshness with "
            "README §7.2 triggers (error expansion / range miss / data stale; "
            "SHADOW-001/WP9-A)"
        ),
    )
    shadow_monitor_cmd.add_argument(
        "--window-days", type=int, default=30, help="recent error window in days (default 30)"
    )
    shadow_monitor_cmd.add_argument(
        "--baseline-ape-median",
        type=float,
        default=0.078,
        help=("error baseline: window APE median above this flags error expansion (G3 §6: 7.8%%)"),
    )
    shadow_monitor_cmd.add_argument(
        "--baseline-range-coverage",
        type=float,
        default=0.80,
        help=("range target: window coverage below this flags range miss (G3 §6: 80-90%%)"),
    )
    shadow_monitor_cmd.add_argument(
        "--stale-days", type=int, default=30, help="data-stale threshold in days (default 30)"
    )
    shadow_monitor_cmd.add_argument("--as-of", help="monitoring date YYYY-MM-DD (default: today)")
    shadow_monitor_cmd.add_argument("--data-dir", type=Path)
    shadow_monitor_cmd.set_defaults(func=_cmd_shadow_monitor)
    shadow_monitor_cmd.set_defaults(shadow_command="monitor")
    shadow_cmd.set_defaults(func=_cmd_shadow_monitor)

    ingest_cmd = subparsers.add_parser(
        "ingest", help="import local raw evidence into immutable parquet snapshots"
    )
    ingest_subparsers = ingest_cmd.add_subparsers(dest="ingest_command")
    file_cmd = ingest_subparsers.add_parser(
        "file", help="import a single local structured file as a raw snapshot"
    )
    file_cmd.add_argument("--input", required=True, type=Path, help="local raw file to import")
    file_cmd.add_argument(
        "--source",
        required=True,
        help="registered source id (e.g. SRC-007) or lake directory name (e.g. lianjia)",
    )
    file_cmd.add_argument("--dataset", required=True, help="dataset name (e.g. chengjiao_list)")
    file_cmd.add_argument(
        "--fetched-at", help="UTC acquisition stamp YYYYMMDD (default: file mtime)"
    )
    file_cmd.add_argument("--query", help="provenance query/URL recorded in the manifest")
    file_cmd.add_argument("--data-dir", type=Path)
    file_cmd.set_defaults(func=_cmd_ingest_file)
    file_cmd.set_defaults(ingest_command="file")
    ingest_cmd.set_defaults(func=_cmd_ingest_file)

    binary_cmd = ingest_subparsers.add_parser(
        "binary",
        help=(
            "import a local raw byte file as an immutable binary snapshot "
            "(EXTFP1-B; data.bin + manifest + RawSnapshot with mime_type)"
        ),
    )
    binary_cmd.add_argument("--input", required=True, type=Path, help="local raw byte file")
    binary_cmd.add_argument(
        "--source",
        required=True,
        help="registered source id (e.g. SRC-011) or lake directory name (e.g. lianjia_ext)",
    )
    binary_cmd.add_argument("--dataset", required=True, help="dataset name (e.g. chengjiao_xlsx)")
    binary_cmd.add_argument(
        "--fetched-at", help="UTC acquisition stamp YYYYMMDD (default: file mtime)"
    )
    binary_cmd.add_argument("--query", help="provenance query/URL recorded in the manifest")
    binary_cmd.add_argument("--mime-type", help="explicit MIME type; default inferred from suffix")
    binary_cmd.add_argument("--data-dir", type=Path)
    binary_cmd.set_defaults(func=_cmd_ingest_binary)
    binary_cmd.set_defaults(ingest_command="binary")

    xlsx_cmd = subparsers.add_parser(
        "xlsx",
        help="外部链家成交 Excel 画像（EXTFP0-D/E；只读，不修改原始文件）",
    )
    xlsx_subparsers = xlsx_cmd.add_subparsers(dest="xlsx_command")
    xlsx_profile_cmd = xlsx_subparsers.add_parser(
        "profile",
        help=(
            "profile 用地分布/面积字段(17/39/40)/描述列/缺失语义，生成机器画像"
            "JSON（只读，绝不修改源 XLSX）"
        ),
    )
    xlsx_profile_cmd.add_argument(
        "--input", required=True, type=Path, help="原始外部链家成交 Excel 路径"
    )
    xlsx_profile_cmd.add_argument(
        "--out", required=True, type=Path, help="机器画像报告 JSON 输出路径"
    )
    xlsx_profile_cmd.set_defaults(func=_cmd_xlsx_profile)
    xlsx_profile_cmd.set_defaults(xlsx_command="profile")

    xlsx_floorplan_cmd = xlsx_subparsers.add_parser(
        "floorplan",
        help=(
            "floorplan 户型图 URL 列表安全解析 + dituFindHouse 占位识别，冻结"
            "普通住宅选择规则并输出机器占位图画像 JSON（只读，不触发下载/OCR）"
        ),
    )
    xlsx_floorplan_cmd.add_argument(
        "--input", required=True, type=Path, help="原始外部链家成交 Excel 路径"
    )
    xlsx_floorplan_cmd.add_argument(
        "--out", required=True, type=Path, help="占位图画像与选择规则 JSON 输出路径"
    )
    xlsx_floorplan_cmd.set_defaults(func=_cmd_xlsx_floorplan)
    xlsx_floorplan_cmd.set_defaults(xlsx_command="floorplan")

    xlsx_parse_cmd = xlsx_subparsers.add_parser(
        "parse",
        help=(
            "parse 全量逐行解析：§6 字段映射 + 缺失语义 + 户型图 URL 衔接，"
            "输出机器解析摘要 JSON（只读，不修改源 XLSX）"
        ),
    )
    xlsx_parse_cmd.add_argument(
        "--input", required=True, type=Path, help="原始外部链家成交 Excel 路径"
    )
    xlsx_parse_cmd.add_argument(
        "--out", required=True, type=Path, help="机器解析摘要 JSON 输出路径"
    )
    xlsx_parse_cmd.add_argument(
        "--sample", type=int, default=0, help="抽样保留前 N 条解析记录到 JSON（默认 0）"
    )
    xlsx_parse_cmd.set_defaults(func=_cmd_xlsx_parse)
    xlsx_parse_cmd.set_defaults(xlsx_command="parse")

    xlsx_stage_cmd = xlsx_subparsers.add_parser(
        "stage",
        help=(
            "stage 全量解析落 staged 表：lianjia_ext_sale_record（全量）+ "
            "lianjia_ext_ordinary_residential（普通住宅）+ 血缘（只读源 XLSX）"
        ),
    )
    xlsx_stage_cmd.add_argument(
        "--input", required=True, type=Path, help="原始外部链家成交 Excel 路径"
    )
    xlsx_stage_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    xlsx_stage_cmd.set_defaults(func=_cmd_xlsx_stage)
    xlsx_stage_cmd.set_defaults(xlsx_command="stage")

    xlsx_quality_cmd = xlsx_subparsers.add_parser(
        "quality",
        help=(
            "quality 从数据湖 staged 两表生成质量报告（守恒/字段质量/URL 分布/"
            "面积一致性 + 回滚点），输出 MD + JSON（只读）"
        ),
    )
    xlsx_quality_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    xlsx_quality_cmd.add_argument(
        "--out", required=True, type=Path, help="机器质量报告 JSON 输出路径"
    )
    xlsx_quality_cmd.add_argument("--git-baseline", help="git 基线 commit（默认取当前 HEAD）")
    xlsx_quality_cmd.set_defaults(func=_cmd_xlsx_quality)
    xlsx_quality_cmd.set_defaults(xlsx_command="quality")

    xlsx_attributes_stage_cmd = xlsx_subparsers.add_parser(
        "attributes-stage",
        help=(
            "attributes-stage 属性标准化 v2 run：从指定 staged run 只读派生"
            "五列标准化属性（楼层/朝向覆盖/电梯/装修/建成年代），不可变新 run "
            "+ 质量摘要 + 指针切换（excel-attribute-enrichment）"
        ),
    )
    xlsx_attributes_stage_cmd.add_argument(
        "--source-run-id",
        required=True,
        help="源 staged run_id（如 20260825T033835Z；只读，不改动）",
    )
    xlsx_attributes_stage_cmd.add_argument(
        "--target-run-id", help="目标 run_id（默认取当前 UTC 时间戳）"
    )
    xlsx_attributes_stage_cmd.add_argument(
        "--data-dir", type=Path, help="数据湖目录（默认 config）"
    )
    xlsx_attributes_stage_cmd.set_defaults(func=_cmd_xlsx_attributes_stage)
    xlsx_attributes_stage_cmd.set_defaults(xlsx_command="attributes-stage")

    # F1（RV-EXTFP0-D-01 积压）收敛：compsval xlsx 不带子命令时给出用法，而非 AttributeError
    xlsx_cmd.set_defaults(func=_cmd_xlsx_unrecognized)

    floorplan_cmd = subparsers.add_parser(
        "floorplan",
        help=(
            "户型图选择清单生成（EXTFP2-B）：从 staged 普通住宅表生成不可变 "
            "selection_manifest（只读，不下载、不触网）"
        ),
    )
    floorplan_subparsers = floorplan_cmd.add_subparsers(dest="floorplan_command")
    floorplan_select_cmd = floorplan_subparsers.add_parser(
        "select",
        help=(
            "select 生成不可变户型图选择清单（普通住宅+URLS_OK+候选>=1→白名单资产；"
            "输出 selection_manifest.json，原子写入）"
        ),
    )
    floorplan_select_cmd.add_argument(
        "--profile",
        default=PROFILE_ORDINARY_RESIDENTIAL_LATEST,
        help=(
            "选择画像（ordinary-residential-latest=全量清单；"
            "production-supported-window-v1=EXTFP4 生产子集；"
            "supported-community-fullhistory-v1=EXTFP6 全历史生产子集）"
        ),
    )
    floorplan_select_cmd.add_argument(
        "--out",
        type=Path,
        help=(
            "选择清单 JSON 输出路径（默认 "
            "data/selection/lianjia_ext/floorplan/selection_manifest.json）"
        ),
    )
    floorplan_select_cmd.add_argument(
        "--run-id", help="覆盖当前 run（默认读 staged/lianjia_ext/current.json 指针）"
    )
    floorplan_select_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    floorplan_select_cmd.set_defaults(func=_cmd_floorplan_select)
    floorplan_select_cmd.set_defaults(floorplan_command="select")

    floorplan_acceptance_cmd = floorplan_subparsers.add_parser(
        "acceptance",
        help=(
            "acceptance 生成 300 张确定性分层验收抽样清单（EXTFP3-G）：区县×年份×居室 "
            "按比例保底抽样，输出与 SelectionManifest 兼容清单，可被 floorplan download 消费"
        ),
    )
    floorplan_acceptance_cmd.add_argument(
        "--profile",
        default=PROFILE_ORDINARY_RESIDENTIAL_LATEST,
        help="选择画像（当前仅 ordinary-residential-latest）",
    )
    floorplan_acceptance_cmd.add_argument(
        "--target",
        type=int,
        default=DEFAULT_TARGET_SIZE,
        help=f"目标抽样记录数（默认 {DEFAULT_TARGET_SIZE}）",
    )
    floorplan_acceptance_cmd.add_argument(
        "--seed",
        type=int,
        default=DEFAULT_SAMPLE_SEED,
        help=f"随机种子（默认 {DEFAULT_SAMPLE_SEED}，与 EXTFP3-G 冻结锚点一致）",
    )
    floorplan_acceptance_cmd.add_argument(
        "--out",
        type=Path,
        help=(
            "抽样清单 JSON 输出路径（默认 "
            "data/selection/lianjia_ext/floorplan/acceptance_manifest.json）"
        ),
    )
    floorplan_acceptance_cmd.add_argument(
        "--golden-csv",
        type=Path,
        help="黄金标签人工标注 CSV 模板输出路径（默认不生成）",
    )
    floorplan_acceptance_cmd.add_argument(
        "--golden-validate",
        type=Path,
        help="校验已标注的黄金标签 CSV（对照 --out 清单）；校验失败返回退出码 2",
    )
    floorplan_acceptance_cmd.add_argument(
        "--run-id", help="覆盖当前 run（默认读 staged/lianjia_ext/current.json 指针）"
    )
    floorplan_acceptance_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    floorplan_acceptance_cmd.set_defaults(func=_cmd_floorplan_acceptance)
    floorplan_acceptance_cmd.set_defaults(floorplan_command="acceptance")

    floorplan_download_cmd = floorplan_subparsers.add_parser(
        "download",
        help=(
            "download 户型图下载器与状态机（EXTFP2-C）：按白名单/幂等键/并发/重试/"
            "断点续跑下载原始字节到 {out}/{download_task_id}.img，维护 download_state.json"
        ),
    )
    floorplan_download_cmd.add_argument(
        "--selection", type=Path, required=True, help="selection_manifest JSON 路径"
    )
    floorplan_download_cmd.add_argument(
        "--out",
        type=Path,
        help=(
            "输出目录（默认 data/download/lianjia_ext/floorplan/），其中维护 download_state.json"
        ),
    )
    floorplan_download_cmd.add_argument(
        "--max-concurrency", type=int, default=4, help="有界并发数（默认 4）"
    )
    floorplan_download_cmd.add_argument(
        "--retries", type=int, default=3, help="单任务最大请求次数（含首次，默认 3）"
    )
    floorplan_download_cmd.add_argument(
        "--timeout", type=float, default=30.0, help="单请求超时秒（默认 30）"
    )
    floorplan_download_cmd.add_argument(
        "--force-new-run",
        action="store_true",
        help="产生新 run_id 与新输出目录，不覆盖旧运行证据",
    )
    floorplan_download_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    floorplan_download_cmd.set_defaults(func=_cmd_floorplan_download)
    floorplan_download_cmd.set_defaults(floorplan_command="download")

    floorplan_asset_cmd = floorplan_subparsers.add_parser(
        "asset",
        help=(
            "asset 原图资产化与字节校验（EXTFP2-D）：读下载运行，校验魔数/MIME/尺寸、"
            "扩展名由字节决定、SHA256 复核、重复标记不合并，落盘原图+生成不可变 manifest"
        ),
    )
    floorplan_asset_cmd.add_argument(
        "--run", type=Path, required=True, help="compsval floorplan download 的输出目录"
    )
    floorplan_asset_cmd.add_argument(
        "--out",
        type=Path,
        help=("数据湖原图目录（默认 data/raw/source=lianjia_ext/dataset=floorplan_image）"),
    )
    floorplan_asset_cmd.add_argument(
        "--batch-id", default="", help="固定 batch_id（默认由下载 run_id 派生）"
    )
    floorplan_asset_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    floorplan_asset_cmd.set_defaults(func=_cmd_floorplan_asset)
    floorplan_asset_cmd.set_defaults(floorplan_command="asset")

    floorplan_e2e_cmd = floorplan_subparsers.add_parser(
        "e2e",
        help=(
            "e2e 样本只读登记 + 10 张真实试跑完整链路（EXTFP2-E）：重建子集清单→真实下载→"
            "原图资产与校验→下载质量报告/链路证据。仅此命令触发真实网络下载（白名单 fail-closed）"
        ),
    )
    floorplan_e2e_cmd.add_argument(
        "--selection",
        type=Path,
        required=True,
        help="全量 selection_manifest JSON（compsval floorplan select 输出）",
    )
    floorplan_e2e_cmd.add_argument(
        "--sample-dir",
        type=Path,
        required=True,
        help="本地 10 张样本目录（含 样本来源清单.md 与 huxingtu_0*.jpg）",
    )
    floorplan_e2e_cmd.add_argument(
        "--sample-list", type=Path, help="样本来源清单.md 路径（默认 --sample-dir/样本来源清单.md）"
    )
    floorplan_e2e_cmd.add_argument(
        "--out", type=Path, help="e2e 证据输出目录（默认 data/selection/lianjia_ext/floorplan/e2e）"
    )
    floorplan_e2e_cmd.add_argument(
        "--expected-count",
        type=int,
        default=10,
        help=(
            "e2e 子集目标资产基数（默认 10）=「seed 命中 + 按稳定顺序补选」的下限；"
            "补选后仍不足即 fail-closed 拒绝生成子集清单（CX-EXTFP2-01 §5）"
        ),
    )
    floorplan_e2e_cmd.add_argument(
        "--max-concurrency", type=int, default=2, help="下载并发上限（默认 2）"
    )
    floorplan_e2e_cmd.add_argument(
        "--retries", type=int, default=3, help="单任务最大请求次数（含首次，默认 3）"
    )
    floorplan_e2e_cmd.add_argument(
        "--timeout", type=float, default=30.0, help="单请求超时秒（默认 30）"
    )
    floorplan_e2e_cmd.add_argument(
        "--force-new-run",
        action="store_true",
        help="产生新下载 run_id 与输出目录，不覆盖旧运行证据",
    )
    floorplan_e2e_cmd.add_argument(
        "--batch-id", default="", help="固定资产 batch_id（默认由下载 run_id 派生）"
    )
    floorplan_e2e_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    floorplan_e2e_cmd.set_defaults(func=_cmd_floorplan_e2e)
    floorplan_e2e_cmd.set_defaults(floorplan_command="e2e")

    floorplan_ocr_cmd = floorplan_subparsers.add_parser(
        "ocr",
        help=(
            "ocr Qwen OCR 请求器与运行记录（EXTFP3-B）：按冻结合同调用 Qwen OCR（网络=是、"
            "付费=是，需授权），逐张保存原始响应与 floorplan_ocr_run 运行记录；幂等键断点续跑"
            "+ 成本门禁 fail-closed"
        ),
    )
    floorplan_ocr_cmd.add_argument(
        "--asset-manifest",
        type=Path,
        required=True,
        help="资产 manifest JSON（compsval floorplan asset 输出的 floorplan_asset_manifest.json）",
    )
    floorplan_ocr_cmd.add_argument(
        "--out",
        type=Path,
        help=(
            "OCR 运行输出根目录（默认 data/raw/source=lianjia_ext/dataset=floorplan_ocr_run，"
            "其下 run_<id>/ 保存运行记录与原始响应）"
        ),
    )
    floorplan_ocr_cmd.add_argument(
        "--config",
        type=Path,
        help="OCR 运行配置 YAML（缺省用冻结合同 + 授权成本上限）",
    )
    floorplan_ocr_cmd.add_argument(
        "--retries", type=int, default=3, help="单图最大请求次数（含首次，默认 3）"
    )
    floorplan_ocr_cmd.add_argument(
        "--concurrency",
        type=int,
        choices=(1, 4, 8, 16),
        default=1,
        help="在飞请求并发数（OCRNEXT-B，默认 1=串行兼容；4 仅诊断档位，验证与验收只走 1/8/16）",
    )
    floorplan_ocr_cmd.add_argument(
        "--timeout", type=float, default=60.0, help="单请求超时秒（默认 60）"
    )
    floorplan_ocr_cmd.add_argument(
        "--force-new-run",
        action="store_true",
        help="产生新 run_id 与新输出目录，不覆盖旧运行证据",
    )
    floorplan_ocr_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    floorplan_ocr_cmd.set_defaults(func=_cmd_floorplan_ocr)
    floorplan_ocr_cmd.set_defaults(floorplan_command="ocr")

    floorplan_transcribe_cmd = floorplan_subparsers.add_parser(
        "transcribe",
        help=(
            "transcribe 确定性转录（EXTFP3-D）：读 OCR 逐词表 → 房间标注表 + 参与字段回填"
            "（纯函数，不触网、不付费）"
        ),
    )
    floorplan_transcribe_cmd.add_argument(
        "--run", type=Path, required=True, help="OCR 运行目录（含 ocr_run.json）"
    )
    floorplan_transcribe_cmd.add_argument(
        "--word-table", type=Path, help="逐词表显式路径（默认 run 目录或 data-dir/staged）"
    )
    floorplan_transcribe_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    floorplan_transcribe_cmd.set_defaults(func=_cmd_floorplan_transcribe)
    floorplan_transcribe_cmd.set_defaults(floorplan_command="transcribe")

    floorplan_verify_cmd = floorplan_subparsers.add_parser(
        "verify",
        help=(
            "verify 自动一致性检查 + 质量报告（EXTFP3-E）：对一次 OCR 运行执行 §9.5 全部检查"
            "并生成 §14 质量报告 JSON；可选回填标注表 consistency_status（只追加派生字段）"
        ),
    )
    floorplan_verify_cmd.add_argument(
        "--run", type=Path, required=True, help="OCR 运行目录（含 ocr_run.json）"
    )
    floorplan_verify_cmd.add_argument(
        "--word-table", type=Path, help="逐词表显式路径（默认自动解析）"
    )
    floorplan_verify_cmd.add_argument(
        "--annotation-table", type=Path, help="标注表显式路径（默认自动解析）"
    )
    floorplan_verify_cmd.add_argument(
        "--asset-manifest", type=Path, help="资产 manifest（floorplan_asset_manifest.json）"
    )
    floorplan_verify_cmd.add_argument(
        "--staged-table", type=Path, help="staged 普通住宅表（Excel 侧比对字段）"
    )
    floorplan_verify_cmd.add_argument(
        "--repeat-annotations", type=Path, help="第二次运行标注表（重复运行一致率）"
    )
    floorplan_verify_cmd.add_argument(
        "--write-consistency",
        action="store_true",
        help="回填标注表 consistency_status（仅追加派生字段，不覆盖原始标注）",
    )
    floorplan_verify_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    floorplan_verify_cmd.set_defaults(func=_cmd_floorplan_verify)
    floorplan_verify_cmd.set_defaults(floorplan_command="verify")

    floorplan_status_cmd = floorplan_subparsers.add_parser(
        "status",
        help=(
            "status 状态机聚合与失败分类（EXTFP3-E §12）：按状态机聚合状态计数，显式分类"
            "失败（可重试/可追溯），供调试与断点续跑"
        ),
    )
    floorplan_status_cmd.add_argument(
        "--run", type=Path, required=True, help="OCR 运行目录（含 ocr_run.json）"
    )
    floorplan_status_cmd.set_defaults(func=_cmd_floorplan_status)
    floorplan_status_cmd.set_defaults(floorplan_command="status")

    floorplan_freeze_cmd = floorplan_subparsers.add_parser(
        "freeze",
        help=(
            "freeze 数据冻结登记（EXTFP5）：只读盘点 lianjia_ext 录入线产物 → 生成 "
            "freeze manifest → 只读校验 → 写版本指针与冻结报告（离线、零付费、不改写既有产物）"
        ),
    )
    floorplan_freeze_cmd.add_argument(
        "--verify-only",
        type=Path,
        help="只校验既有 freeze manifest（只读重算哈希/计数比对，不写任何文件）",
    )
    floorplan_freeze_cmd.add_argument("--data-dir", type=Path, help="数据湖目录（默认 config）")
    floorplan_freeze_cmd.set_defaults(func=_cmd_floorplan_freeze)
    floorplan_freeze_cmd.set_defaults(floorplan_command="freeze")
    floorplan_cmd.set_defaults(func=_cmd_floorplan_unrecognized)

    sql_cmd = subparsers.add_parser(
        "sql", help="run SQL against the latest raw snapshots (views named raw_<dataset>)"
    )
    sql_cmd.add_argument("query")
    sql_cmd.add_argument("--max-rows", type=int, default=40)
    sql_cmd.add_argument("--data-dir", type=Path)
    sql_cmd.set_defaults(func=_cmd_sql)

    data_cmd = subparsers.add_parser(
        "data",
        help="run the reproducible stage/mart + quality-report pipeline",
    )
    data_subparsers = data_cmd.add_subparsers(dest="data_command")
    stage_cmd = data_subparsers.add_parser(
        "stage",
        help="re-derive staged + marts tables and the quality report from a snapshot",
    )
    stage_cmd.add_argument(
        "--snapshot", required=True, help="snapshot id (e.g. lianjia/chengjiao@20260821T000000Z)"
    )
    stage_cmd.add_argument("--data-dir", type=Path)
    stage_cmd.set_defaults(func=_cmd_data_stage)
    stage_cmd.set_defaults(data_command="stage")
    marts_build_cmd = data_subparsers.add_parser(
        "marts-build",
        help="build the combined multi-source marts (lianjia + fang_esf) with cross-source dedup",
    )
    marts_build_cmd.add_argument("--data-dir", type=Path)
    marts_build_cmd.set_defaults(func=_cmd_data_marts_build)
    marts_build_cmd.set_defaults(data_command="marts-build")
    marts_enrich_cmd = data_subparsers.add_parser(
        "marts-enrich-attributes",
        help=(
            "身份键 join 把 staged v2 标准化属性回填 valid_sale（显式 --out，"
            "不改写既有 mart；excel-attribute-enrichment）"
        ),
    )
    marts_enrich_cmd.add_argument(
        "--run-id", help="staged v2 run_id（默认取 staged current 指针）"
    )
    marts_enrich_cmd.add_argument(
        "--out", required=True, type=Path, help="扩列 valid_sale 输出路径"
    )
    marts_enrich_cmd.add_argument("--data-dir", type=Path)
    marts_enrich_cmd.set_defaults(func=_cmd_data_marts_enrich)
    marts_enrich_cmd.set_defaults(data_command="marts-enrich-attributes")
    data_cmd.set_defaults(func=_cmd_data_stage)

    entities_cmd = subparsers.add_parser(
        "entities",
        help="build WP5 entity authority tables (community, alias, building, market_series, scope)",
    )
    entities_subparsers = entities_cmd.add_subparsers(dest="entities_command")
    entities_build_cmd = entities_subparsers.add_parser(
        "build",
        help=(
            "build WP5 entity tables from the candidate 名录 skeleton and print the "
            "待人工确认 conflict checklist + scope policy list"
        ),
    )
    entities_build_cmd.add_argument("--data-dir", type=Path)
    entities_build_cmd.set_defaults(func=_cmd_entities_build)
    entities_build_cmd.set_defaults(entities_command="build")
    entities_cmd.set_defaults(func=_cmd_entities_build)

    valuation_cmd = subparsers.add_parser(
        "valuation",
        help="build WP6 valuation intermediate tables (candidate pool)",
    )
    valuation_subparsers = valuation_cmd.add_subparsers(dest="valuation_command")
    valuation_build_cmd = valuation_subparsers.add_parser(
        "build",
        help=(
            "build subject_property/valuation_run/comp_candidate from a subject "
            "JSON (WP6-A candidate pool, VAL1-002)"
        ),
    )
    valuation_build_cmd.add_argument(
        "--subject",
        required=True,
        type=Path,
        help="subject JSON file (数据字典 §3.9)",
    )
    valuation_build_cmd.add_argument("--data-dir", type=Path)
    valuation_build_cmd.set_defaults(func=_cmd_valuation_build)
    valuation_build_cmd.set_defaults(valuation_command="build")
    valuation_tier_cmd = valuation_subparsers.add_parser(
        "tier",
        help=(
            "fill comparable tier/similarity into comp_candidate and write "
            "candidate competitive-community relations (WP6-B, VAL1-003)"
        ),
    )
    valuation_tier_cmd.add_argument(
        "--subject",
        required=True,
        type=Path,
        help="subject JSON file (数据字典 §3.9)",
    )
    valuation_tier_cmd.add_argument("--data-dir", type=Path)
    valuation_tier_cmd.set_defaults(func=_cmd_valuation_tier)
    valuation_tier_cmd.set_defaults(valuation_command="tier")
    valuation_time_cmd = valuation_subparsers.add_parser(
        "time",
        help=(
            "apply time adjustment to comparable candidates and write "
            "comp_adjustment (adjustment_type=时间) (WP6-C, VAL1-004)"
        ),
    )
    valuation_time_cmd.add_argument(
        "--subject",
        required=True,
        type=Path,
        help="subject JSON file (数据字典 §3.9)",
    )
    valuation_time_cmd.add_argument("--data-dir", type=Path)
    valuation_time_cmd.set_defaults(func=_cmd_valuation_time)
    valuation_time_cmd.set_defaults(valuation_command="time")
    valuation_diff_cmd = valuation_subparsers.add_parser(
        "diff",
        help=(
            "apply property difference adjustment to comparable candidates and "
            "write comp_adjustment (adjustment_type=差异) (WP6-D, VAL1-005)"
        ),
    )
    valuation_diff_cmd.add_argument(
        "--subject",
        required=True,
        type=Path,
        help="subject JSON file (数据字典 §3.9)",
    )
    valuation_diff_cmd.add_argument("--data-dir", type=Path)
    valuation_diff_cmd.set_defaults(func=_cmd_valuation_diff)
    valuation_diff_cmd.set_defaults(valuation_command="diff")
    valuation_aggregate_cmd = valuation_subparsers.add_parser(
        "aggregate",
        help=(
            "aggregate center/range/confidence/output-status and write "
            "valuation_result (WP6-E, VAL1-006)"
        ),
    )
    valuation_aggregate_cmd.add_argument(
        "--subject",
        required=True,
        type=Path,
        help="subject JSON file (数据字典 §3.9)",
    )
    valuation_aggregate_cmd.add_argument("--data-dir", type=Path)
    valuation_aggregate_cmd.set_defaults(func=_cmd_valuation_aggregate)
    valuation_aggregate_cmd.set_defaults(valuation_command="aggregate")
    valuation_review_cmd = valuation_subparsers.add_parser(
        "review",
        help=(
            "append a human review event: write review_event (append-only, "
            "never overwrite automatic result) (WP6-F, VAL1-007)"
        ),
    )
    valuation_review_cmd.add_argument(
        "--input",
        required=True,
        type=Path,
        help="review JSON contract file (技术方案 §11.1)",
    )
    valuation_review_cmd.add_argument("--data-dir", type=Path)
    valuation_review_cmd.set_defaults(func=_cmd_valuation_review)
    valuation_review_cmd.set_defaults(valuation_command="review")
    valuation_cmd.set_defaults(func=_cmd_valuation_build)

    system_cmd = subparsers.add_parser(
        "system",
        help="run the offline quality gate (ruff + mypy + pytest)",
    )
    system_cmd.add_subparsers(dest="system_command").add_parser(
        "check", help="run the offline quality gate"
    ).set_defaults(func=_cmd_system_check)
    system_cmd.set_defaults(func=_cmd_system_check)
    subparsers.add_parser("system-check", help="alias for `compsval system check`").set_defaults(
        func=_cmd_system_check
    )

    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    func = getattr(args, "func", None)
    if func is None:
        parser.print_help()
        return 2
    return int(func(args))


if __name__ == "__main__":
    raise SystemExit(main())
