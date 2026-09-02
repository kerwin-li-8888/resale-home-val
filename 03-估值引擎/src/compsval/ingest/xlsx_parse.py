"""外部链家成交 Excel 全量逐行解析（EXTFP1-C，技术方案 §6）。

把外部链家成交 Excel 逐行解析为标准化记录（``XlsxParsedRecord``），落实
技术方案 §6 的字段映射与缺失语义（§6.3/§6.4），并衔接 EXTFP0-E 的户型图
URL 安全解析（§3.2）。本模块只读，绝不修改原始 XLSX。

输入支持两种等价形态（输出同一解析结果）：

- 原始 XLSX（``parse_xlsx``）：openpyxl 只读流式读取，按表头名匹配列
  （不硬编码列号，列顺序变化时仍可复核）；
- cell 级结构化快照（``transpose_cell_table``）：把 EXTFP0-C 的
  ``sheet_name/row/column/value/value_type`` 五列 cell 表转置回
  ``(header, rows)``，再走同一解析逻辑（§13 原始层 cell 表 → 逐行解析结果）。

关键口径（2026-08-24 探针实测，见 §9 EXTFP1-C 记录）：

- ``成交总价``：字符串十进制数字，单位**元**（``700000/32.09≈21814`` 与
  ``成交均价`` 一致验证）；
- ``成交均价``：字符串十进制数字，单位 元/㎡（来源披露值，保留原文）；
- ``挂牌价格``：字符串十进制数字，单位**万元**（解析为元时 ×10000）；
- ``成交日期``：openpyxl 原生 ``datetime``（data_only=False 下为 datetime 对象）。

缺失语义（§6.4）：``暂无/暂无数据/空`` → ``MISSING``（来源未披露）；
原文存在但无法解析 → ``PARSE_FAILURE``；解析成功 → ``PARSED``。解析失败
绝不自动变成 0 或空串。
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from hashlib import sha256
from pathlib import Path
from typing import Any

from pydantic import BaseModel, Field

from compsval.ingest.floorplan_profile import (
    UrlListStatus,
    parse_url_list,
)

# 解析规则版本，随摘要报告一起落盘（可追溯）
PARSE_RULE_VERSION = "EXTFP1-C-1.0"

# 表头常量（技术方案 §6；与 profile_xlsx/floorplan_profile 保持一致）
HEADER_SOURCE_RECORD_ID = "房屋ID"
HEADER_COMMUNITY_NAME = "小区名字"
HEADER_COMMUNITY_SOURCE_ID = "小区ID"
HEADER_SALE_DATE = "成交日期"
HEADER_TOTAL_PRICE = "成交总价"
HEADER_UNIT_PRICE = "成交均价"
HEADER_AREA_17 = "房屋面积"
HEADER_AREA_39 = "房屋面积.1"
HEADER_AREA_40 = "建筑面积"
HEADER_LAYOUT = "户型"
HEADER_BEDROOMS = "卧室数量"
HEADER_LIVING_ROOMS = "客厅数量"
HEADER_FLOOR = "楼层"
HEADER_ORIENTATION = "朝向"
HEADER_DECORATION = "装修情况"
HEADER_BUILT_YEAR = "建成时间"
HEADER_HOUSE_TYPE = "房屋类型"
HEADER_HAS_ELEVATOR = "是否有电梯"
HEADER_PROPERTY_USE = "房屋用途"
HEADER_FLOORPLAN_URL = "户型图"
HEADER_LISTING_PRICE = "挂牌价格"
HEADER_LISTING_DAYS = "成交天数"
HEADER_PRICE_ADJUSTMENTS = "价格调整次数"
HEADER_DESC = "房源描述"
HEADER_DESC_TAGS = "房源描述.1"

# 「来源未披露/暂无」同义占位（§6.4 MISSING）
NON_ASSERTED = {"暂无", "暂无数据"}
# 面积字段可能带单位后缀（EXTFP0-D 探针确认含 "㎡"）
AREA_UNIT_SUFFIXES = ("㎡", "m²", "m2", "平方米", "平米")


class FieldParseStatus(StrEnum):
    """数值/日期字段的解析状态（§6.4 缺失语义）。"""

    PARSED = "PARSED"
    MISSING = "MISSING"  # 空 / 暂无类占位（来源未披露）
    PARSE_FAILURE = "PARSE_FAILURE"  # 原文存在但无法解析


class PropertyUseNorm(StrEnum):
    """房屋用途归一（§4.1：仅 普通住宅 进入本轮结构化范围）。"""

    ORDINARY_RESIDENTIAL = "普通住宅"
    ASSERTED_EXCLUDED = "非普通住宅"  # 商住两用/车库/别墅/平房等明确排除
    UNKNOWN = "UNKNOWN"  # 空 / 暂无 / 暂无数据


class XlsxParsedRecord(BaseModel):
    """单行标准化解析记录（技术方案 §6.3 最小成交映射 + 来源扩展字段）。"""

    row_number: int = Field(description="数据行号（表头行=0，数据从 1 开始）；raw_locator 锚点")
    source_record_id: str = Field(description="房屋ID 原文（保留字符串原值）")
    community_name: str = Field(description="来源小区名（不合并到现有小区实体）")
    community_source_id: str = Field(description="来源小区ID（不合并到现有小区实体）")
    sale_date_raw: str = Field(description="成交日期 原文（datetime 转 ISO 字符串）")
    sale_date: str | None = Field(default=None, description="解析日期 YYYY-MM-DD；未知=None")
    sale_date_precision: str = Field(default="UNKNOWN", description="日期精度 DAY/MONTH/UNKNOWN")
    total_price_raw: str = Field(description="成交总价 原文（保留字符串）")
    total_price_yuan: Decimal | None = Field(
        default=None, gt=0, description="成交总价（元，单位已验证）；未知=None 不用 0"
    )
    total_price_status: FieldParseStatus = Field(description="总价解析状态")
    unit_price_observed_raw: str = Field(description="成交均价 原文（来源披露值）")
    unit_price_observed: Decimal | None = Field(
        default=None, gt=0, description="成交均价（元/㎡，来源披露）；未知=None 不用 0"
    )
    unit_price_status: FieldParseStatus = Field(description="均价解析状态")
    transaction_area_sqm: Decimal | None = Field(
        default=None, gt=0, description="第17列 房屋面积（㎡）；未知=None 不用 0"
    )
    area_status: FieldParseStatus = Field(description="第17列面积解析状态")
    building_area_detail_sqm: Decimal | None = Field(
        default=None, gt=0, description="第40列 建筑面积（㎡，独立来源字段）；未知=None"
    )
    building_area_status: FieldParseStatus = Field(description="第40列面积解析状态")
    layout_raw: str = Field(description="户型 原文（保留原文和标准值）")
    bedrooms_raw: str = Field(description="卧室数量 原文")
    living_rooms_raw: str = Field(description="客厅数量 原文")
    floor_raw: str = Field(description="楼层 原文（如 中楼层/8层）")
    orientation: str = Field(description="朝向 原文")
    decoration: str = Field(description="装修情况 原文")
    built_year_raw: str = Field(description="建成时间 原文（如 2015年/暂无数据）")
    house_type: str = Field(description="房屋类型 原文（塔楼/板楼…）")
    has_elevator_raw: str = Field(description="是否有电梯 原文")
    property_use_raw: str = Field(description="房屋用途 原文")
    property_use_norm: PropertyUseNorm = Field(description="用途归一")
    floorplan_url_list_raw: str | None = Field(
        default=None, description="户型图 原文列表字符串（完整保留，不截断）"
    )
    floorplan_url_status: str = Field(
        default=UrlListStatus.NO_URL.value, description="NO_URL/URL_PARSE_FAILURE/URLS_OK"
    )
    floorplan_candidate_count: int = Field(
        default=0, description="户型图候选 URL 数（非 dituFindHouse 占位）"
    )
    listing_price_raw: str = Field(description="挂牌价格 原文（单位万元，解析为元时 ×10000）")
    listing_price_yuan: Decimal | None = Field(
        default=None, gt=0, description="挂牌价（元）；未知=None 不用 0"
    )
    listing_price_status: FieldParseStatus = Field(description="挂牌价解析状态")
    listing_days_raw: str = Field(description="成交天数 原文")
    listing_days: int | None = Field(default=None, description="成交天数；未知=None")
    listing_days_status: FieldParseStatus = Field(description="成交天数解析状态")
    price_adjustments_raw: str = Field(description="价格调整次数 原文")
    source_property_description: str = Field(description="第7列 房源描述（独立保留）")
    source_property_tags: str = Field(description="第19列 房源描述.1（独立保留）")
    extra_fields: dict[str, str] = Field(
        default_factory=dict, description="其余列原文全量保留（来源扩展字段）"
    )


@dataclass(frozen=True)
class ParseSummary:
    """全量解析摘要（机器产物，随报告落盘）。"""

    data_rows_total: int
    parsed_count: int
    property_use_distribution: dict[str, int]
    ordinary_residential_count: int
    field_status_counts: dict[str, dict[str, int]]
    source_sha256: str | None = None
    parse_rule_version: str = PARSE_RULE_VERSION


@dataclass
class ParseResult:
    """解析结果：记录（惰性迭代）+ 摘要。"""

    records: Iterator[XlsxParsedRecord]
    summary: ParseSummary


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _is_non_asserted(value: Any) -> bool:
    return isinstance(value, str) and value.strip() in NON_ASSERTED


def _text(value: Any) -> str:
    """单元格 → 保留原文的字符串（datetime 用 str() 与 cell 表 value 形态一致）。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return str(value)
    if isinstance(value, bool):
        return "True" if value else "False"
    return str(value)


def _strip_area_unit(text: str) -> str:
    for suffix in AREA_UNIT_SUFFIXES:
        if text.endswith(suffix):
            return text[: -len(suffix)].strip()
    return text


def _parse_decimal(text: str, *, wan_to_yuan: bool = False) -> Decimal | None:
    """把字符串解析为 Decimal；不可解析返回 None（由调用方定 PARSE_FAILURE）。

    ``wan_to_yuan``：挂牌价口径为万元，×10000 得元。
    """
    try:
        value = Decimal(text.replace(",", ""))
    except InvalidOperation:
        return None
    if wan_to_yuan:
        value *= Decimal("10000")
    return value


def _parse_positive_decimal_field(raw: Any, *, wan_to_yuan: bool = False) -> tuple[
    Decimal | None, FieldParseStatus
]:
    """数值字段统一解析：空/暂无→MISSING；原文无法解析→PARSE_FAILURE；成功→PARSED。"""
    if _is_empty(raw):
        return None, FieldParseStatus.MISSING
    text = _text(raw).strip()
    if _is_non_asserted(text):
        return None, FieldParseStatus.MISSING
    value = _parse_decimal(text, wan_to_yuan=wan_to_yuan)
    if value is None or value <= 0:
        return None, FieldParseStatus.PARSE_FAILURE
    return value, FieldParseStatus.PARSED


def _parse_area_field(raw: Any) -> tuple[Decimal | None, FieldParseStatus]:
    """面积字段：先剥离单位后缀再解析（EXTFP0-D 修复口径）。"""
    if _is_empty(raw):
        return None, FieldParseStatus.MISSING
    text = _text(raw).strip()
    if _is_non_asserted(text):
        return None, FieldParseStatus.MISSING
    value = _parse_decimal(_strip_area_unit(text))
    if value is None or value <= 0:
        return None, FieldParseStatus.PARSE_FAILURE
    return value, FieldParseStatus.PARSED


def _parse_sale_date(raw: Any) -> tuple[str, str | None, str]:
    """成交日期：datetime → ISO 日期（DAY）；字符串按 ISO/数字解析；失败 → 保留原文+UNKNOWN。"""
    original = _text(raw)
    if isinstance(raw, datetime):
        return original, raw.date().isoformat(), "DAY"
    if isinstance(raw, str) and raw.strip():
        text = raw.strip()
        for fmt in (
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d",
            "%Y.%m.%d",
        ):
            try:
                dt = datetime.strptime(text, fmt)
                return original, dt.date().isoformat(), "DAY"
            except ValueError:
                continue
    return original, None, "UNKNOWN"


def _norm_use(raw: Any) -> tuple[str, PropertyUseNorm]:
    """房屋用途归一：普通住宅 / 明确排除 / UNKNOWN（空/暂无）。"""
    if _is_empty(raw) or _is_non_asserted(raw):
        return "", PropertyUseNorm.UNKNOWN
    norm = str(raw).strip()
    if norm == PropertyUseNorm.ORDINARY_RESIDENTIAL.value:
        return norm, PropertyUseNorm.ORDINARY_RESIDENTIAL
    return norm, PropertyUseNorm.ASSERTED_EXCLUDED


def parse_row(row_number: int, row: tuple[Any, ...], idx: dict[str, int]) -> XlsxParsedRecord:
    """把一行原始单元格解析为标准化记录（纯函数，便于单测与 cell 表输入复用）。"""
    len_row = len(row)

    def at(header: str) -> Any:
        i = idx.get(header)
        return row[i] if (i is not None and i < len_row) else None

    def at_text(header: str) -> str:
        return _text(at(header))

    total_price, total_status = _parse_positive_decimal_field(at(HEADER_TOTAL_PRICE))
    unit_price, unit_status = _parse_positive_decimal_field(at(HEADER_UNIT_PRICE))
    area17, area17_status = _parse_area_field(at(HEADER_AREA_17))
    area40, area40_status = _parse_area_field(at(HEADER_AREA_40))
    listing_price, listing_status = _parse_positive_decimal_field(
        at(HEADER_LISTING_PRICE), wan_to_yuan=True
    )
    listing_days, days_status = _parse_positive_decimal_field(at(HEADER_LISTING_DAYS))
    sale_raw, sale_date, sale_precision = _parse_sale_date(at(HEADER_SALE_DATE))
    use_raw, use_norm = _norm_use(at(HEADER_PROPERTY_USE))

    # 户型图 URL：原文完整保留 + 安全解析衔接（EXTFP0-E）
    floorplan_raw = at(HEADER_FLOORPLAN_URL)
    floorplan_text = _text(floorplan_raw).strip() or None
    url_status: UrlListStatus
    items: list[Any]
    if floorplan_text is None or _is_non_asserted(floorplan_text):
        url_status, items = UrlListStatus.NO_URL, []
    else:
        url_status, items = parse_url_list(floorplan_text)
    candidate_count = sum(1 for it in items if it.url_class.value == "FLOORPLAN_CANDIDATE")

    # 其余列全量保留为来源扩展字段（原始层不丢信息）
    extra: dict[str, str] = {}
    for name, i in idx.items():
        if i is not None and i < len_row and name not in _EXTRACTED_HEADERS:
            extra[name] = _text(row[i])

    return XlsxParsedRecord(
        row_number=row_number,
        source_record_id=at_text(HEADER_SOURCE_RECORD_ID),
        community_name=at_text(HEADER_COMMUNITY_NAME),
        community_source_id=at_text(HEADER_COMMUNITY_SOURCE_ID),
        sale_date_raw=sale_raw,
        sale_date=sale_date,
        sale_date_precision=sale_precision,
        total_price_raw=at_text(HEADER_TOTAL_PRICE),
        total_price_yuan=total_price,
        total_price_status=total_status,
        unit_price_observed_raw=at_text(HEADER_UNIT_PRICE),
        unit_price_observed=unit_price,
        unit_price_status=unit_status,
        transaction_area_sqm=area17,
        area_status=area17_status,
        building_area_detail_sqm=area40,
        building_area_status=area40_status,
        layout_raw=at_text(HEADER_LAYOUT),
        bedrooms_raw=at_text(HEADER_BEDROOMS),
        living_rooms_raw=at_text(HEADER_LIVING_ROOMS),
        floor_raw=at_text(HEADER_FLOOR),
        orientation=at_text(HEADER_ORIENTATION),
        decoration=at_text(HEADER_DECORATION),
        built_year_raw=at_text(HEADER_BUILT_YEAR),
        house_type=at_text(HEADER_HOUSE_TYPE),
        has_elevator_raw=at_text(HEADER_HAS_ELEVATOR),
        property_use_raw=use_raw,
        property_use_norm=use_norm,
        floorplan_url_list_raw=floorplan_text,
        floorplan_url_status=url_status.value,
        floorplan_candidate_count=candidate_count,
        listing_price_raw=at_text(HEADER_LISTING_PRICE),
        listing_price_yuan=listing_price,
        listing_price_status=listing_status,
        listing_days_raw=at_text(HEADER_LISTING_DAYS),
        listing_days=int(listing_days) if listing_days is not None else None,
        listing_days_status=days_status,
        price_adjustments_raw=at_text(HEADER_PRICE_ADJUSTMENTS),
        source_property_description=at_text(HEADER_DESC),
        source_property_tags=at_text(HEADER_DESC_TAGS),
        extra_fields=extra,
    )


# 已显式映射/提取的表头（不重复进入 extra_fields）
_EXTRACTED_HEADERS = frozenset(
    {
        HEADER_SOURCE_RECORD_ID,
        HEADER_COMMUNITY_NAME,
        HEADER_COMMUNITY_SOURCE_ID,
        HEADER_SALE_DATE,
        HEADER_TOTAL_PRICE,
        HEADER_UNIT_PRICE,
        HEADER_AREA_17,
        HEADER_AREA_40,
        HEADER_LAYOUT,
        HEADER_BEDROOMS,
        HEADER_LIVING_ROOMS,
        HEADER_FLOOR,
        HEADER_ORIENTATION,
        HEADER_DECORATION,
        HEADER_BUILT_YEAR,
        HEADER_HOUSE_TYPE,
        HEADER_HAS_ELEVATOR,
        HEADER_PROPERTY_USE,
        HEADER_FLOORPLAN_URL,
        HEADER_LISTING_PRICE,
        HEADER_LISTING_DAYS,
        HEADER_PRICE_ADJUSTMENTS,
        HEADER_DESC,
        HEADER_DESC_TAGS,
    }
)


def _sheet_sha256(path: Path) -> str:
    hasher = sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _parse_rows(
    header: tuple[Any, ...], rows: Iterator[tuple[Any, ...]]
) -> Iterator[XlsxParsedRecord]:
    idx: dict[str, int] = {}
    for i, v in enumerate(header):
        name = str(v) if v is not None else ""
        if name and name not in idx:
            idx[name] = i
    for row_number, row in enumerate(rows, start=1):
        yield parse_row(row_number, tuple(row), idx)


def iter_parse_xlsx(path: Path) -> Iterator[XlsxParsedRecord]:
    """流式解析 XLSX（生成器；openpyxl 只读，绝不修改源文件）。

    迭代器耗尽或调用方显式 ``close()`` 时自动关闭 workbook。每个产出的记录
    为 ``XlsxParsedRecord``；真实文件 245,410 行，调用方应流式消费（CLI 只
    累计摘要，不保留全量列表，避免内存放大）。
    """
    from openpyxl import load_workbook

    if not path.is_file():
        raise FileNotFoundError(f"source xlsx not found: {path}")
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            header = ()
        yield from _parse_rows(header, rows)
    finally:
        wb.close()


def parse_xlsx(path: Path) -> ParseResult:
    """流式解析 XLSX 的入口：返回记录迭代器 + 占位摘要（含源 SHA256）。

    调用方迭代记录后用 :func:`summarize` 生成最终摘要（守恒/分布/缺失统计）。
    """
    return ParseResult(
        records=iter_parse_xlsx(path),
        summary=ParseSummary(
            data_rows_total=0,
            parsed_count=0,
            property_use_distribution={},
            ordinary_residential_count=0,
            field_status_counts={},
            source_sha256=_sheet_sha256(path),
        ),
    )


def summarize(
    records: list[XlsxParsedRecord] | Iterator[XlsxParsedRecord],
    source_sha256: str | None = None,
) -> ParseSummary:
    """对已消费的记录生成解析摘要（守恒/分布/缺失统计）。"""
    use_dist: dict[str, int] = {}
    ordinary = 0
    status_counts: dict[str, dict[str, int]] = {}
    count = 0
    for rec in records:
        count += 1
        use_dist[rec.property_use_norm.value] = use_dist.get(rec.property_use_norm.value, 0) + 1
        if rec.property_use_norm is PropertyUseNorm.ORDINARY_RESIDENTIAL:
            ordinary += 1
        for field_name, status in (
            ("total_price", rec.total_price_status),
            ("unit_price", rec.unit_price_status),
            ("area_17", rec.area_status),
            ("area_40", rec.building_area_status),
            ("listing_price", rec.listing_price_status),
            ("listing_days", rec.listing_days_status),
            ("sale_date", FieldParseStatus.PARSED if rec.sale_date else (
                FieldParseStatus.MISSING
                if not rec.sale_date_raw
                else FieldParseStatus.PARSE_FAILURE
            )),
        ):
            bucket = status_counts.setdefault(field_name, {})
            bucket[status.value] = bucket.get(status.value, 0) + 1
    return ParseSummary(
        data_rows_total=count,
        parsed_count=count,
        property_use_distribution=use_dist,
        ordinary_residential_count=ordinary,
        field_status_counts=status_counts,
        source_sha256=source_sha256,
    )


def _col_letter_to_index(letter: str) -> int:
    """Excel 列字母 → 数值序号（A=1, Z=26, AA=27…），用于 cell 表转置排序。"""
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def transpose_cell_table(
    table: Any,
) -> dict[str, tuple[tuple[str, ...], list[tuple[int, list[Any]]]]]:
    """把 cell 级结构化快照（sheet_name/row/column/value/value_type）转置回行。

    按 sheet 名分组，返回 ``{sheet_name: (header, rows)}``：``header`` 为表头名
    元组（row=1 的单元格，按 Excel 列字母数值序排列），``rows`` 为
    ``(row_number, values)`` 列表（values 与 header 对齐，cell 表中不存在的
    单元格用 None 填充）。与 ``iter_parse_xlsx`` 的 openpyxl 行形态等价，可喂给
    ``_parse_rows`` 得到同一解析结果（§13 原始层 cell 表 → 逐行解析结果）。
    """
    sheets: dict[str, dict[int, dict[str, str]]] = {}
    for row, sheet_name, col, val in zip(
        table.column("row").to_pylist(),
        table.column("sheet_name").to_pylist(),
        table.column("column").to_pylist(),
        table.column("value").to_pylist(),
        strict=True,
    ):
        sheets.setdefault(str(sheet_name), {}).setdefault(int(row), {})[str(col)] = str(val)

    result: dict[str, tuple[tuple[str, ...], list[tuple[int, list[Any]]]]] = {}
    for sheet_name, entries in sheets.items():
        header_row = entries.get(1, {})
        column_letters = sorted(header_row, key=_col_letter_to_index)
        header = tuple(header_row[c] for c in column_letters)
        rows: list[tuple[int, list[Any]]] = []
        for row_number in sorted(k for k in entries if k != 1):
            values = [entries.get(row_number, {}).get(c) for c in column_letters]
            rows.append((row_number, values))
        result[sheet_name] = (header, rows)
    return result


__all__ = [
    "FieldParseStatus",
    "PARSE_RULE_VERSION",
    "ParseResult",
    "ParseSummary",
    "PropertyUseNorm",
    "XlsxParsedRecord",
    "iter_parse_xlsx",
    "parse_row",
    "parse_xlsx",
    "summarize",
    "transpose_cell_table",
]
