"""Import a local structured raw file into an immutable raw parquet snapshot.

WP4-A (DATA-004): ``compsval ingest file --input <path> --source <src> --dataset
<name>`` reads a local structured file and persists it as an immutable raw
snapshot (parquet + manifest) using the WP1 ``write_raw_snapshot`` primitive.
No domain interpretation happens here:

- ``.csv`` files are parsed structurally with PyArrow's CSV reader, preserving
  columns and types as observed.
- ``.txt`` / other text files are preserved one row per byte-exact line
  (``line_no``, ``content``), so the evidence bytes remain recoverable and the
  WP4-B source parser can re-read them losslessly.

Importing the same file to the same ``fetched_at`` stamp raises
``FileExistsError`` from ``write_raw_snapshot`` (single-snapshot semantics):
a re-run never overwrites existing evidence.
"""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import pyarrow as pa
import pyarrow.csv as pacsv
from openpyxl import load_workbook

from compsval.contract.registry import SOURCE_ID_BY_DIR
from compsval.ingest.snapshots import SnapshotResult, write_raw_snapshot


def resolve_source_dir(token: str) -> str:
    """Map a ``--source`` token to a lake partition directory name.

    Accepts either a registered source id (``SRC-007``) or the directory name
    itself (``lianjia``). Unknown tokens are rejected so typos do not silently
    create a partition that ``compsval catalog`` cannot relate to a registry source.
    """
    if token in SOURCE_ID_BY_DIR:
        return token
    for dirname, source_id in SOURCE_ID_BY_DIR.items():
        if source_id == token:
            return dirname
    known = ", ".join(sorted(set(SOURCE_ID_BY_DIR) | set(SOURCE_ID_BY_DIR.values())))
    raise ValueError(f"unknown --source {token!r}; expected one of: {known}")


def _value_type(value: Any) -> str:
    """Classify an openpyxl cell value into a coarse Arrow-friendly type tag.

    Materializing a tag alongside the stringified value lets a later parser
    distinguish numbers/bools/dates from strings without re-sniffing text.
    """
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        return "string"
    if isinstance(value, (datetime, date, time)):
        return "datetime"
    return "unknown"


def read_xlsx_table(path: Path) -> pa.Table:
    """Read an ``.xlsx`` workbook structurally into a raw PyArrow table.

    EXTFP0-C (技术方案 §5.2/§8.5): XLSX is read via openpyxl's read-only
    streaming mode, never faked as UTF-8 text and never re-saved as CSV (which
    would lose the original evidence). Each populated cell becomes one row with
    ``sheet_name``/``row``/``column``/``value``/``value_type`` so the raw content
    stays recoverable and type-tagged without assuming a header layout.
    """
    sheet_names: list[str] = []
    rows: list[int] = []
    columns: list[str] = []
    values: list[str] = []
    types: list[str] = []
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    sheet_names.append(ws.title)
                    rows.append(cell.row)
                    columns.append(cell.column_letter)
                    values.append(str(cell.value))
                    types.append(_value_type(cell.value))
    finally:
        wb.close()
    return pa.table(
        {
            "sheet_name": pa.array(sheet_names, type=pa.string()),
            "row": pa.array(rows, type=pa.int64()),
            "column": pa.array(columns, type=pa.string()),
            "value": pa.array(values, type=pa.string()),
            "value_type": pa.array(types, type=pa.string()),
        }
    )


def file_to_table(path: Path) -> pa.Table:
    """Read a local structured file into a raw PyArrow table.

    CSV is parsed structurally; ``.xlsx`` workbooks are read structurally via
    :func:`read_xlsx_table` (openpyxl read-only, never faked as UTF-8 text);
    every other extension is treated as raw text and stored one line per row
    (``line_no`` 1-based, ``content``), dropping a single trailing newline so an
    empty trailing row is not recorded.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pacsv.read_csv(path)
    if suffix == ".xlsx":
        return read_xlsx_table(path)
    lines = path.read_text(encoding="utf-8").split("\n")
    if lines and lines[-1] == "":
        lines = lines[:-1]
    return pa.table(
        {
            "line_no": pa.array(range(1, len(lines) + 1), type=pa.int64()),
            "content": pa.array(lines, type=pa.string()),
        }
    )


def import_local_file(
    *,
    input_path: Path,
    source: str,
    dataset: str,
    fetched_at: datetime,
    query: str,
    endpoint: str = "",
    data_dir: Path | None = None,
) -> SnapshotResult:
    """Import a local structured file as an immutable raw snapshot.

    ``source`` may be a registry source id (``SRC-007``) or a lake directory
    name (``lianjia``); it is canonicalized through :func:`resolve_source_dir`.
    ``query`` records how the file was obtained (URL or local path) in the
    manifest provenance.
    """
    source_dir = resolve_source_dir(source)
    table = file_to_table(input_path)
    return write_raw_snapshot(
        table,
        root=data_dir,
        source=source_dir,
        dataset=dataset,
        fetched_at=fetched_at,
        query=query,
        endpoint=endpoint,
        source_row_count=table.num_rows,
    )