"""普通住宅画像与字段映射（EXTFP0-D）。

对外部链家成交 Excel（``01-数据/外部数据/示例城市 (2).xlsx``）做只读画像，
输出机器可解析的 JSON 画像报告。本模块只读不写，绝不修改原始 XLSX。
统计目标对应技术方案 §3/§6：

- 用途分布（第 24 列 ``房屋用途``：普通住宅 / 商住两用 / 车库 / 别墅 / 平房 / 空值）；
- 面积字段（第 17/39/40 列：``房屋面积`` / ``房屋面积.1`` / ``建筑面积``）
  的可解析性、范围与两两一致性（§6.1）；
- 描述列（第 7 列 ``房源描述`` 与第 19 列 ``房源描述.1``）的分布（§6.2）；
- 缺失语义（UNKNOWN / MISSING / PARSE_FAILURE），存在不可解析与来源未披露严格区分（§6.4）。

技术方案 §3.1 的数字是方案制定时的只读画像，不构成实施证据；本报告是正式实施
工作包的机器画像产物，由本脚本对真实文件重跑生成。

字段按表头名匹配并记录实际所在列（字母/序号），不硬编码列号，列顺序变化时
画像仍可复核而不是静默错位。
"""

from __future__ import annotations

from collections import Counter
from datetime import UTC, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

# 技术方案 §6.1 固定的目标表头 → (期望位置说明, 标准化口径)
AREA_TARGETS: tuple[tuple[str, str], ...] = (
    ("房屋面积", "第17列 Q 房屋面积 → transaction_area_sqm"),
    ("房屋面积.1", "第39列 AM 房屋面积.1（与第17列重复，不进入标准化业务表）"),
    ("建筑面积", "第40列 AN 建筑面积（独立来源字段 building_area_detail_sqm）"),
)

USE_HEADER = "房屋用途"
DESC_PAIR_HEADERS = ("房源描述", "房源描述.1")

# 「来源未披露 / 暂无」同义占位归一到同一桶
USE_NON_ASSERTED = {"暂无", "暂无数据"}

# 面积字段可能带单位后缀（外部链家样本经探针确认含 "㎡"）
AREA_UNIT_SUFFIXES = ("㎡", "m²", "m2", "平方米", "平米")

# 画像规则版本，随报告一起落盘，便于追溯
PROFILE_RULE_VERSION = "EXTFP0-D-1.0"


class HeaderInfo(BaseModel):
    """表头信息：位置 + 名称，供画像复核列映射是否正确。"""

    column_index: int
    column_letter: str
    name: str


class AreaColumnStats(BaseModel):
    """单个面积字段的画像。"""

    header: str
    column_index: int
    column_letter: str
    mapping: str
    data_rows_total: int
    parseable_count: int
    parse_failure_count: int = Field(description="原文存在但无法解析为数值")
    missing_count: int = Field(description="空/缺失（来源未披露）")
    min_value: float | None
    max_value: float | None
    sum_value: float | None
    count_gt_zero: int
    equal_to_17_count: int | None = Field(
        default=None, description="与房屋面积列（第17列）两者都可解析且一致的行数"
    )
    differs_from_17_count: int | None = Field(
        default=None, description="与房屋面积列两者都可解析但不一致的行数"
    )
    both_parseable_count: int | None = Field(
        default=None, description="与房屋面积列两者都可解析的行数"
    )


class DescriptionPairStats(BaseModel):
    """描述列（第7列/第19列）画像。"""

    source_column_index: int
    source_header: str
    tags_column_index: int
    tags_header: str
    source_present_count: int
    tags_present_count: int
    both_empty_count: int
    both_present_same_count: int
    both_present_differ_count: int


class MissingSemantics(BaseModel):
    """单列缺失语义画像（技术方案 §6.4）。"""

    header: str
    column_index: int
    column_letter: str
    missing_count: int
    non_assert_placeholder_count: int = Field(
        default=0, description="值为 暂无/暂无数据 类占位的行数"
    )


class SheetProfile(BaseModel):
    """单个工作表的画像。"""

    sheet_name: str
    data_rows_total: int
    columns_total: int
    headers: list[HeaderInfo]
    property_use_distribution: dict[str, int]
    ordinary_residential_count: int
    property_use_asserted_excluded: int = Field(
        description="用途明确但非普通住宅（商住两用/车库/别墅/平房等）行数"
    )
    property_use_unknown: int = Field(
        description="用途缺失或占位（空/暂无/暂无数据）行数"
    )
    area_columns: list[AreaColumnStats]
    description_pair: DescriptionPairStats | None
    missing_semantics: list[MissingSemantics]


class XlsxProfileReport(BaseModel):
    """机器画像报告（EXTFP0-D 退出证据）。"""

    source_path: str
    source_sha256: str | None = None
    profile_rule_version: str = PROFILE_RULE_VERSION
    profiled_at: str | None = None
    sheets: list[SheetProfile]


def _is_empty(value: Any) -> bool:
    """单元格是否视为空（None 或去除空白后的空串）。"""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _at(row: tuple[Any, ...], len_row: int, c: int | None) -> Any:
    """安全取行内第 c 个单元格；越界或 None 一律返回 None。"""
    return row[c] if (c is not None and c < len_row) else None


def _to_number(value: Any) -> float | None:
    """尽力把单元格解析为数值；布尔、空白、来源未披露占位返回 None。

    先剥离面积单位后缀（如 ``㎡``），再尝试解析为 float。无法解析（如 ``-``）
    返回 None，由调用方计入 parse_failure。
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "":
        return None
    for suffix in AREA_UNIT_SUFFIXES:
        if text.endswith(suffix):
            text = text[: -len(suffix)].strip()
            break
    try:
        return float(text)
    except ValueError:
        try:
            return float(text.replace(",", ""))
        except ValueError:
            return None


def _is_non_assert_placeholder(value: Any) -> bool:
    """是否为「来源未披露/暂无」同义占位（技术方案 §6.4 的 MISSING 语义）。"""
    return isinstance(value, str) and value.strip() in USE_NON_ASSERTED


def _sheet_sha256(path: Path) -> str:
    """全文件 SHA256（血缘：报告客观对应某一次原始证据字节）。"""
    hasher = sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def profile_xlsx(path: Path, out_json: Path | None = None) -> XlsxProfileReport:
    """流式读取 XLSX 生成机器画像报告；绝不修改源文件。

    :param path: 原始 Excel 路径（只读）。
    :param out_json: 非空时以原子方式写入 JSON 报告（tmp + replace）。
    """
    if not path.is_file():
        raise FileNotFoundError(f"source xlsx not found: {path}")

    sheets: list[SheetProfile] = []
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        for ws in wb.worksheets:
            sheets.append(_profile_sheet(ws))
    finally:
        wb.close()

    report = XlsxProfileReport(
        source_path=str(path.resolve()),
        source_sha256=_sheet_sha256(path),
        profile_rule_version=PROFILE_RULE_VERSION,
        profiled_at=datetime.now(UTC).isoformat(),
        sheets=sheets,
    )

    if out_json is not None:
        out_json.parent.mkdir(parents=True, exist_ok=True)
        work = out_json.with_name(out_json.name + ".incomplete")
        work.write_text(report.model_dump_json(indent=2), encoding="utf-8")
        work.replace(out_json)

    return report


def _profile_sheet(ws: Any) -> SheetProfile:
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        header = ()

    columns_total = len(header)
    header_infos = [
        HeaderInfo(
            column_index=i,
            column_letter=get_column_letter(i + 1),
            name=str(v) if v is not None else "",
        )
        for i, v in enumerate(header)
    ]

    index_by_header: dict[str, int] = {}
    for info in header_infos:
        if info.name and info.name not in index_by_header:
            index_by_header[info.name] = info.column_index

    idx = index_by_header  # 别名

    use_idx = idx.get(USE_HEADER)
    area_idxs: list[int | None] = [idx.get(t) for t, _ in AREA_TARGETS]
    desc_idxs: list[int | None] = [idx.get(h) for h in DESC_PAIR_HEADERS]

    missing_specs: list[tuple[str, int | None]] = [
        (t, idx.get(t)) for t, _ in AREA_TARGETS
    ] + [(h, idx.get(h)) for h in DESC_PAIR_HEADERS]

    use_dist: Counter[str] = Counter()
    ordinary = 0
    use_excluded = 0
    use_unknown = 0

    # 面积累计：仅对表头存在的列累计解析/缺失/失败
    present_area_idxs = [i for i in area_idxs if i is not None]
    area_min: dict[int, float] = {}
    area_max: dict[int, float] = {}
    area_sum: dict[int, float] = {}
    area_parseable: dict[int, int] = {}
    area_failure: dict[int, int] = {}
    area_missing: dict[int, int] = {}
    area_gt0: dict[int, int] = {}
    for i in present_area_idxs:
        area_min[i] = 0.0
        area_max[i] = 0.0
        area_sum[i] = 0.0
        area_parseable[i] = 0
        area_failure[i] = 0
        area_missing[i] = 0
        area_gt0[i] = 0

    # 与第17列一致性
    col17 = area_idxs[0]
    cons17: dict[int, tuple[int, int]] = {i: (0, 0) for i in present_area_idxs[1:]}

    desc15: dict[str, int] = {
        "source_present": 0,
        "tags_present": 0,
        "both_empty": 0,
        "both_same": 0,
        "both_differ": 0,
    }

    missing_counter: dict[tuple[str, int | None], tuple[int, int]] = {}
    for key in missing_specs:
        missing_counter[key] = (0, 0)

    data_rows = 0
    for row in rows:
        data_rows += 1
        len_row = len(row)

        # 用途
        if use_idx is not None:
            use_val = _at(row, len_row, use_idx)
            empty = _is_empty(use_val)
            if isinstance(use_val, str):
                norm = use_val.strip()
            elif use_val is None:
                norm = None
            else:
                norm = str(use_val)
            if empty or norm is None:
                bucket = "(空)"
            elif norm in USE_NON_ASSERTED:
                bucket = norm
            else:
                bucket = norm
            use_dist[bucket] += 1
            if not empty and norm is not None and norm not in USE_NON_ASSERTED:
                if norm == "普通住宅":
                    ordinary += 1
                else:
                    use_excluded += 1
            else:
                use_unknown += 1

        # 面积字段（先收集再比较，保证一个行内三列一致性）
        vals: dict[int, float | None] = {}
        for i in present_area_idxs:
            raw = _at(row, len_row, i)
            if _is_empty(raw) or _is_non_assert_placeholder(raw):
                # 空单元格与「暂无/暂无数据」同属 来源未披露（MISSING），非解析失败
                area_missing[i] += 1
                vals[i] = None
                continue
            num = _to_number(raw)
            if num is None:
                area_failure[i] += 1
                vals[i] = None
                continue
            area_parseable[i] += 1
            area_sum[i] += num
            if num > 0:
                area_gt0[i] += 1
            area_min[i] = num if area_parseable[i] == 1 else min(area_min[i], num)
            area_max[i] = num if area_parseable[i] == 1 else max(area_max[i], num)
            vals[i] = num
        # 与第17列一致性：要求第17列与比较列两列都可解析才计入（RV-EXTFP0-D-03#F6）
        if col17 is not None and col17 in vals and vals[col17] is not None:
            v17 = vals[col17]
            for i in present_area_idxs[1:]:
                vi = vals.get(i)
                if vi is not None:
                    eq, diff = cons17[i]
                    if vi == v17:
                        eq += 1
                    else:
                        diff += 1
                    cons17[i] = (eq, diff)

        # 描述对
        s_val = _at(row, len_row, desc_idxs[0])
        t_val = _at(row, len_row, desc_idxs[1])
        s_present = not _is_empty(s_val)
        t_present = not _is_empty(t_val)
        if s_present:
            desc15["source_present"] += 1
        if t_present:
            desc15["tags_present"] += 1
        if not s_present and not t_present:
            desc15["both_empty"] += 1
        elif s_present and t_present:
            if str(s_val) == str(t_val):
                desc15["both_same"] += 1
            else:
                desc15["both_differ"] += 1

        # 缺失语义
        for (h, c) in missing_specs:
            raw = _at(row, len_row, c)
            m, p = missing_counter[(h, c)]
            if _is_empty(raw):
                m += 1
            elif _is_non_assert_placeholder(raw):
                p += 1
            missing_counter[(h, c)] = (m, p)

    # 组装面积字段画像（顺序对齐 AREA_TARGETS）
    area_columns: list[AreaColumnStats] = []
    for pos, (target, note) in enumerate(AREA_TARGETS):
        col_idx: int | None = area_idxs[pos]
        if col_idx is None:
            area_columns.append(
                AreaColumnStats(
                    header=target,
                    column_index=-1,
                    column_letter="",
                    mapping=note,
                    data_rows_total=data_rows,
                    parseable_count=0,
                    parse_failure_count=0,
                    missing_count=0,
                    min_value=None,
                    max_value=None,
                    sum_value=None,
                    count_gt_zero=0,
                    equal_to_17_count=None,
                    differs_from_17_count=None,
                    both_parseable_count=None,
                )
            )
            continue
        eq, diff = cons17.get(col_idx, (0, 0))
        area_columns.append(
            AreaColumnStats(
                header=target,
                column_index=col_idx,
                column_letter=get_column_letter(col_idx + 1),
                mapping=note,
                data_rows_total=data_rows,
                parseable_count=area_parseable[col_idx],
                parse_failure_count=area_failure[col_idx],
                missing_count=area_missing[col_idx],
                min_value=area_min[col_idx] if area_parseable[col_idx] else None,
                max_value=area_max[col_idx] if area_parseable[col_idx] else None,
                sum_value=area_sum[col_idx] if area_parseable[col_idx] else None,
                count_gt_zero=area_gt0[col_idx],
                equal_to_17_count=eq if pos > 0 else None,
                differs_from_17_count=diff if pos > 0 else None,
                both_parseable_count=(eq + diff) if pos > 0 else None,
            )
        )

    # 描述对
    if desc_idxs[0] is not None and desc_idxs[1] is not None:
        description_pair = DescriptionPairStats(
            source_column_index=desc_idxs[0],
            source_header=DESC_PAIR_HEADERS[0],
            tags_column_index=desc_idxs[1],
            tags_header=DESC_PAIR_HEADERS[1],
            source_present_count=desc15["source_present"],
            tags_present_count=desc15["tags_present"],
            both_empty_count=desc15["both_empty"],
            both_present_same_count=desc15["both_same"],
            both_present_differ_count=desc15["both_differ"],
        )
    else:
        description_pair = None

    missing_semantics = [
        MissingSemantics(
            header=h,
            column_index=c if c is not None else -1,
            column_letter=get_column_letter(c + 1) if c is not None else "",
            missing_count=m,
            non_assert_placeholder_count=p,
        )
        for (h, c), (m, p) in missing_counter.items()
    ]

    return SheetProfile(
        sheet_name=ws.title,
        data_rows_total=data_rows,
        columns_total=columns_total,
        headers=header_infos,
        property_use_distribution=dict(
            sorted(use_dist.items(), key=lambda kv: -kv[1])
        ),
        ordinary_residential_count=ordinary,
        property_use_asserted_excluded=use_excluded,
        property_use_unknown=use_unknown,
        area_columns=area_columns,
        description_pair=description_pair,
        missing_semantics=missing_semantics,
    )