"""占位图画像与选择规则（EXTFP0-E）。

对外部链家成交 Excel 的 ``户型图`` 列做只读画像，并冻结「普通住宅有效户型图候选」
选择规则。本模块只读不写，绝不修改原始 XLSX。

背景（技术方案 §3.2）：``户型图`` 单元格不是单个裸 URL，而是字符串形式的 URL 列表，
例如 ``['http://ke-image.ljcdn.com/hdic-frame/xx.jpg.1440x1080.jpg?from=ke.com']``。
``/beike/dituFindHouse/`` 路径是地图或找房占位资源的强信号，不能直接视为户型图；
仅检查 HTTP 状态码会把这类资源错误计为成功户型图。EXTFP0 在此冻结分类规则，确认前
失败关闭——本包只做离线画像与规则冻结，不下载、不 OCR。

范围边界（EXTFP0 合同 §2）：不下载任何图片、不调用实时网络或付费 API、不写密钥/Cookie；
不跳过登录/验证码/反爬规则；选择清单与下载属 EXTFP1 之后的授权工作包，不在本包。
"""

from __future__ import annotations

import ast
from collections import Counter
from datetime import UTC, datetime
from enum import StrEnum
from hashlib import sha256
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

from pydantic import BaseModel, Field

# 户型图列表头（外部链家成交 Excel）
URL_LIST_HEADER = "户型图"
USE_HEADER = "房屋用途"

# 「来源未披露 / 暂无」同义占位（技术方案 §6.4 MISSING）
NON_ASSERTED = {"暂无", "暂无数据"}

# 地图/找房占位资源强信号路径片段（技术方案 §3.2）
PLACEHOLDER_PATH_SIGNAL = "/dituFindHouse/"

# 冻结的选择规则版本：随报告落盘，供复核与下游（EXTFP1）引用
SELECTION_RULE_VERSION = "EXTFP0-E-SELECT-1.1"

# 选择规则正文（机器可读，随画像报告一起输出，视为冻结版本证据）
SELECTION_RULE_TEXT = "\n".join(
    [
        "[EXTFP0-E-SELECT-1.1] 普通住宅有效户型图候选选择规则",
        "1. 记录必须是 房屋用途=普通住宅；",
        "2. 户型图 必须能安全解析出 >=1 个字符串 URL（空/占位/解析失败/基数异常零条均非候选）；",
        "3. 至少一个 URL 为户型图候选（http(s) 且不含 /beike/dituFindHouse/ 占位信号），"
        "   全为怀疑占位图的记录标记 NON_FLOORPLAN_PLACEHOLDER，不进入选择清单；",
        "4. 多 URL 记录按其 URL 序号分别登记，不静默取第一条；",
        "5. 本规则只做离线画像与候选分类，不触发任何下载/OCR；下载与 OCR 属后续授权工作包。",
    ]
)


class UrlListStatus(StrEnum):
    """``户型图`` 单元格 URL 列表的整记录分类（技术方案 §6.4/§7.5）。"""

    NO_URL = "NO_URL"  # 空 / 暂无类占位（MISSING，来源未披露）
    URL_PARSE_FAILURE = "URL_PARSE_FAILURE"  # 存在原文但非可解析 URL 列表 / 基数异常零条
    URLS_OK = "URLS_OK"  # 解析出 >=1 个字符串 URL


class UrlClass(StrEnum):
    """单个 URL 的分类。"""

    FLOORPLAN_CANDIDATE = "FLOORPLAN_CANDIDATE"  # 户型图候选，可进入选择清单
    PLACEHOLDER = "PLACEHOLDER"  # dituFindHouse 地图/找房占位，不能当户型图


class UrlListItem(BaseModel):
    """解析出的单个 URL 及其分类。"""

    url: str = Field(description="原始 URL，不改写")
    url_class: UrlClass


class FloorplanRecordMetrics(BaseModel):
    """整张表（Sheet1）的户型图画像。"""

    data_rows_total: int
    url_list_status_counts: dict[str, int]  # UrlListStatus 计数，含非普通住宅
    multi_url_records: int = Field(description="URLS_OK 且含 >1 个 URL 的记录数")
    url_total: int = Field(description="解析出的 URL 总数")
    url_class_counts: dict[str, int]  # UrlClass 计数


class OrdinaryFloorplanMetrics(BaseModel):
    """普通住宅子集的户型图选择画像（冻结规则的目标群体）。"""

    ordinary_residential_count: int
    no_url_count: int = Field(description="普通住宅中 户型图 无/占位 记录数")
    parse_failure_count: int = Field(description="普通住宅中 URL 列表解析失败/基数异常记录数")
    placeholder_only_count: int = Field(
        description="普通住宅中全部 URL 为 dituFindHouse 占位（NON_FLOORPLAN_PLACEHOLDER）记录数"
    )
    floorplan_candidate_count: int = Field(
        description="普通住宅中 >=1 个户型图候选 URL（SELECT-1.1 第3条通过）记录数"
    )
    multi_url_candidate_count: int = Field(
        description="普通住宅户型图候选记录中含多 URL 的记录数（供基数异常核实）"
    )
    placeholder_url_count: int = Field(description="普通住宅中占位 URL 总数")
    floorplan_url_count: int = Field(description="普通住宅中户型图候选 URL 总数")


class FloorplanProfileReport(BaseModel):
    """占位图画像与选择规则冻结报告（EXTFP0-E 退出证据）。"""

    source_path: str
    source_sha256: str | None = None
    selection_rule_version: str = SELECTION_RULE_VERSION
    selection_rule_text: str = SELECTION_RULE_TEXT
    column_letter: str = Field(description="户型图 列字母")
    use_column_letter: str = Field(description="房屋用途 列字母")
    profiled_at: str | None = None
    url_list: FloorplanRecordMetrics
    ordinary: OrdinaryFloorplanMetrics


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _is_non_asserted(value: Any) -> bool:
    return isinstance(value, str) and value.strip() in NON_ASSERTED


def _at(row: tuple[Any, ...], len_row: int, c: int | None) -> Any:
    return row[c] if (c is not None and c < len_row) else None


def is_placeholder_url(url: str) -> bool:
    """URL 是否为地图/找房占位资源（技术方案 §3.2）。"""
    try:
        path = urlsplit(url).path or ""
    except ValueError:
        path = url
    return PLACEHOLDER_PATH_SIGNAL in path


def classify_url(url: str) -> UrlClass:
    """单 URL 分类：户型图候选 vs 非候选（RV-EXTFP0-E-03#F2）。

    规则 EXTFP0-E-SELECT-1.1 第 3 条：户型图候选必须是 ``http(s)`` 且不含
    dituFindHouse 占位信号。非 ``http(s)`` scheme 的 URL（如 ``ftp:``/``data:``）
    不是合法户型图候选，一律归 ``PLACEHOLDER``（非候选桶），不进入选择清单。
    """
    try:
        scheme = (urlsplit(url).scheme or " ").lower()
    except ValueError:
        scheme = " "
    if scheme not in ("http", "https"):
        return UrlClass.PLACEHOLDER
    return UrlClass.PLACEHOLDER if is_placeholder_url(url) else UrlClass.FLOORPLAN_CANDIDATE


def parse_url_list(raw: Any) -> tuple[UrlListStatus, list[UrlListItem]]:
    """安全解析 ``户型图`` 单元格（技术方案 §3.2：禁止 eval，用字面量解析）。

    返回 (状态, 解析出的 URL 列表)。解析失败/非列表/零条 URLs 返回 URL_PARSE_FAILURE。
    """
    if _is_empty(raw):
        return UrlListStatus.NO_URL, []
    text = str(raw).strip()
    if text in NON_ASSERTED:
        return UrlListStatus.NO_URL, []
    try:
        obj = ast.literal_eval(text)
    except (ValueError, SyntaxError, MemoryError):
        return UrlListStatus.URL_PARSE_FAILURE, []
    if not isinstance(obj, list):
        return UrlListStatus.URL_PARSE_FAILURE, []
    # 只取字符串元素；非字符串元素整体视为基数异常
    str_urls = []
    for item in obj:
        if isinstance(item, str) and item.strip():
            str_urls.append(item.strip())
        elif isinstance(item, str) and not item.strip():
            continue  # 空串忽略
        else:
            str_urls = []  # 非字符串元素 → 异常
    if not str_urls:
        return UrlListStatus.URL_PARSE_FAILURE, []
    items = [UrlListItem(url=u, url_class=classify_url(u)) for u in str_urls]
    return UrlListStatus.URLS_OK, items


def _sheet_sha256(path: Path) -> str:
    hasher = sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index + 1)


def profile_floorplan(path: Path, out_json: Path | None = None) -> FloorplanProfileReport:
    """流式读取 XLSX 生成占位图画像 + 冻结选择规则报告；绝不修改源文件。"""
    if not path.is_file():
        raise FileNotFoundError(f"source xlsx not found: {path}")

    from openpyxl import load_workbook

    url_status: Counter[str] = Counter()
    url_class_total: Counter[str] = Counter()
    url_total = 0
    multi_url_total = 0

    ordinary_residential = 0
    ordinary_no_url = 0
    ordinary_parse_failure = 0
    ordinary_placeholder_only = 0
    ordinary_candidate = 0
    ordinary_multi_candidate = 0
    ordinary_placeholder_url = 0
    ordinary_floorplan_url = 0

    data_rows = 0
    column_letter = ""
    use_column_letter = ""

    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                header = ()
            index_by_name: dict[str, int] = {}
            for i, v in enumerate(header):
                name = str(v) if v is not None else ""
                if name and name not in index_by_name:
                    index_by_name[name] = i
            url_idx = index_by_name.get(URL_LIST_HEADER)
            use_idx = index_by_name.get(USE_HEADER)
            column_letter = _column_letter(url_idx) if url_idx is not None else ""
            use_column_letter = _column_letter(use_idx) if use_idx is not None else ""

            for row in rows:
                data_rows += 1
                len_row = len(row)
                # 用途判断
                is_ordinary = False
                use_val = _at(row, len_row, use_idx)
                if not _is_empty(use_val) and not _is_non_asserted(use_val):
                    is_ordinary = str(use_val).strip() == "普通住宅"
                if is_ordinary:
                    ordinary_residential += 1

                raw = _at(row, len_row, url_idx)
                status, items = parse_url_list(raw)
                url_status[status.value] += 1
                if status == UrlListStatus.URLS_OK:
                    url_total += len(items)
                    for it in items:
                        url_class_total[it.url_class.value] += 1
                    if len(items) > 1:
                        multi_url_total += 1

                if is_ordinary:
                    if status == UrlListStatus.NO_URL:
                        ordinary_no_url += 1
                    elif status == UrlListStatus.URL_PARSE_FAILURE:
                        ordinary_parse_failure += 1
                    else:
                        placeholders = [it for it in items if it.url_class == UrlClass.PLACEHOLDER]
                        candidates = [
                            it for it in items if it.url_class == UrlClass.FLOORPLAN_CANDIDATE
                        ]
                        ordinary_placeholder_url += len(placeholders)
                        ordinary_floorplan_url += len(candidates)
                        if candidates:
                            ordinary_candidate += 1
                            if len(items) > 1:
                                ordinary_multi_candidate += 1
                        else:
                            ordinary_placeholder_only += 1
    finally:
        wb.close()

    report = FloorplanProfileReport(
        source_path=str(path.resolve()),
        source_sha256=_sheet_sha256(path),
        selection_rule_version=SELECTION_RULE_VERSION,
        selection_rule_text=SELECTION_RULE_TEXT,
        column_letter=column_letter,
        use_column_letter=use_column_letter,
        profiled_at=datetime.now(UTC).isoformat(),
        url_list=FloorplanRecordMetrics(
            data_rows_total=data_rows,
            url_list_status_counts=dict(url_status),
            multi_url_records=multi_url_total,
            url_total=url_total,
            url_class_counts=dict(url_class_total),
        ),
        ordinary=OrdinaryFloorplanMetrics(
            ordinary_residential_count=ordinary_residential,
            no_url_count=ordinary_no_url,
            parse_failure_count=ordinary_parse_failure,
            placeholder_only_count=ordinary_placeholder_only,
            floorplan_candidate_count=ordinary_candidate,
            multi_url_candidate_count=ordinary_multi_candidate,
            placeholder_url_count=ordinary_placeholder_url,
            floorplan_url_count=ordinary_floorplan_url,
        ),
    )

    if out_json is not None:
        out_json.parent.mkdir(parents=True, exist_ok=True)
        work = out_json.with_name(out_json.name + ".incomplete")
        work.write_text(report.model_dump_json(indent=2), encoding="utf-8")
        work.replace(out_json)

    return report
